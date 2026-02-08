import sys
import pandas as pd
import re
import os
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog
from datetime import datetime
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import seaborn as sns
import platform
import numpy as np
import math 
from mpl_toolkits.axes_grid1 import make_axes_locatable
from matplotlib.colors import ListedColormap 
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

class ElectionAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("사전투표운용장비 배분 최적화 시스템")
        self.root.geometry("1100x750") 
        self.root.resizable(True, True) 
        
        self.vote_files = []
        self.cached_data = {} 
        self.equipment_file = None
        self.file_past_elect = None   
        self.file_recent_elect = None 
        
        self.region_name = "" 

        self.last_reserve_count = 5
        self.station_data = {}

        # [추가] 시뮬레이션용 집계 변수
        self.total_past_voters = 0     # 과거 사전투표자 총합 (A)
        self.total_past_intra = 0      # [신규] 과거 관내 총합
        self.total_past_extra = 0      # [신규] 과거 관외 총합
        self.total_recent_electors = 0 # 이번 선거인수 총합 (C)
        self.total_past_electors = 0   # [추가] 과거 선거인수 총합 (분모용)
        self.past_turnout_rate = 0.0   # 과거 사전투표율 (기준)
        
        self.create_widgets()

    def _get_merged_rate_text(self, r_intra, r_extra):
        def _fmt(val):
            # 혹시 모를 문자열 입력 대비
            try:
                val = float(val)
            except:
                return str(val)

            if val > 0: return f"+ {val:.1f}%"   # .1f 추가 (소수점 1자리)
            elif val < 0: return f"- {abs(val):.1f}%" # .1f 추가
            else: return "-"
            
        if r_intra == r_extra:
            return _fmt(r_intra)
        else:
            return f"관내:{_fmt(r_intra)} / 관외:{_fmt(r_extra)}"
            
    def create_widgets(self):
        # [구조 변경] 좌우 2단 분할 레이아웃
        main_container = ttk.Frame(self.root, padding="15")
        main_container.pack(fill="both", expand=True)

        # === [좌측 패널] 컨트롤러 ===
        left_panel = ttk.Frame(main_container, width=320)
        left_panel.pack(side="left", fill="y", expand=False, padx=(0, 15))
        left_panel.pack_propagate(False) 

        # === [우측 패널] 데이터 뷰어 ===
        right_panel = ttk.Frame(main_container)
        right_panel.pack(side="right", fill="both", expand=True)

        # -------------------------------------------------------
        # [좌측 1] 기초 데이터 로드
        # -------------------------------------------------------
        frame_data = ttk.LabelFrame(left_panel, text=" 1. 기초 데이터 로드 ", padding="10")
        frame_data.pack(fill="x", pady=(0, 15))
        
        btn_files = ttk.Button(frame_data, text="📂 시간대별 투표 데이터 파일", command=self.select_vote_files)
        btn_files.pack(fill="x", ipady=5)
        self.lbl_file_count = ttk.Label(frame_data, text="파일 없음", foreground="gray", font=("맑은 고딕", 9))
        self.lbl_file_count.pack(pady=(2, 8))

        btn_equip = ttk.Button(frame_data, text="📂 운용장비 현황 파일", command=self.select_equip_file)
        btn_equip.pack(fill="x", ipady=5)
        self.lbl_equip_status = ttk.Label(frame_data, text="파일 미선택 (기본값: 1대)", foreground="gray", font=("맑은 고딕", 9))
        self.lbl_equip_status.pack(pady=(2, 8))

        frame_elect = ttk.Frame(frame_data)
        frame_elect.pack(fill="x", pady=(5, 0))
        
        # [수정] UI/UX 반영: '현황' 제거하여 가독성 확보
        btn_past = ttk.Button(frame_elect, text="📂 직전 인구수통보", command=self.select_past_file)
        btn_past.pack(side="left", fill="x", expand=True, padx=(0, 2))
        
        # [수정] UI/UX 반영: '현황' 제거하여 가독성 확보
        btn_recent = ttk.Button(frame_elect, text="📂 당해 인구수통보", command=self.select_recent_file)
        btn_recent.pack(side="right", fill="x", expand=True, padx=(2, 0))
        
        self.lbl_elect_status = ttk.Label(frame_data, text="파일 미선택 (변동률 미적용)", foreground="gray", font=("맑은 고딕", 9))
        self.lbl_elect_status.pack(pady=(2, 5))
        
        # -------------------------------------------------------
        # [좌측 2] 사전투표 운용장비 (순서 변경: 3번)
        # -------------------------------------------------------
        frame_equip = ttk.LabelFrame(left_panel, text=" 3. 사전투표 운용장비 ", padding="10")
        frame_equip.pack(fill="x", pady=(0, 15))

        btn_balance = ttk.Button(frame_equip, text="⚖️ 운용장비 자동 배분 실행", command=self.open_balance_popup)
        btn_balance.pack(fill="x", ipady=6)

        # -------------------------------------------------------
        # [좌측 3] 분석 리포트 (순서 변경: 4번 / 옵션 통합)
        # -------------------------------------------------------
        frame_report = ttk.LabelFrame(left_panel, text=" 4. 혼잡도 분석 리포트 ", padding="10")
        frame_report.pack(fill="x", pady=(0, 15))

        # 1) 분석 기준 (라디오 버튼)
        f_mode_label = ttk.Frame(frame_report)
        f_mode_label.pack(fill="x", pady=(0, 2))
        ttk.Label(f_mode_label, text="분석 기준:", font=("맑은 고딕", 9, "bold")).pack(anchor="w")

        self.var_mode = tk.StringVar(value="density")
        f_radio = ttk.Frame(frame_report)
        f_radio.pack(fill="x", pady=(0, 8))
        
        r1 = ttk.Radiobutton(f_radio, text="발급자수(장비1대기준)", variable=self.var_mode, value="density")
        r1.pack(side="left", expand=True, anchor="w")
        r2 = ttk.Radiobutton(f_radio, text="사전투표자 수", variable=self.var_mode, value="population")
        r2.pack(side="left", expand=True, anchor="w")

        ttk.Separator(frame_report, orient="horizontal").pack(fill="x", pady=5)

        # 2) 보기 옵션 (체크박스)
        f_opt_label = ttk.Frame(frame_report)
        f_opt_label.pack(fill="x", pady=(5, 2))
        ttk.Label(f_opt_label, text="보기 옵션:", font=("맑은 고딕", 9, "bold")).pack(anchor="w")

        self.var_day1 = tk.BooleanVar(value=True)
        self.var_day2 = tk.BooleanVar(value=True)
        self.var_intra = tk.BooleanVar(value=True)
        self.var_extra = tk.BooleanVar(value=True)
        self.var_day_all = tk.BooleanVar(value=True) 

        chk_f1 = ttk.Frame(frame_report)
        chk_f1.pack(fill="x", pady=2)
        ttk.Label(chk_f1, text="기간: ").pack(side="left")
        ttk.Checkbutton(chk_f1, text="1일", variable=self.var_day1).pack(side="left", padx=2)
        ttk.Checkbutton(chk_f1, text="2일", variable=self.var_day2).pack(side="left", padx=2)
        ttk.Checkbutton(chk_f1, text="전체", variable=self.var_day_all).pack(side="left", padx=2)
        
        chk_f2 = ttk.Frame(frame_report)
        chk_f2.pack(fill="x", pady=2)
        ttk.Label(chk_f2, text="구분: ").pack(side="left")
        ttk.Checkbutton(chk_f2, text="관내", variable=self.var_intra).pack(side="left", padx=5)
        ttk.Checkbutton(chk_f2, text="관외", variable=self.var_extra).pack(side="left", padx=5)

        # 3) 리포트 출력 버튼
        btn_run = ttk.Button(frame_report, text="🚀 분석 리포트 출력", command=self.run_simulation)
        btn_run.pack(fill="x", ipady=6, pady=(10, 0))

        # [좌측 하단] 초기화 버튼 (맨 아래 고정)
        style = ttk.Style()
        style.configure("Danger.TButton", foreground="red", font=("맑은 고딕", 9))
        btn_reset = ttk.Button(left_panel, text="🗑️ 모든 데이터 초기화 (주의)", command=self.reset_all, style="Danger.TButton")
        btn_reset.pack(side="bottom", fill="x", pady=(10, 0))

        # -------------------------------------------------------
        # [좌측 5] 부가 기능 (통합)
        # -------------------------------------------------------
        frame_sub = ttk.LabelFrame(left_panel, text=" 5. 기표대 및 롤 투표용지 ", padding="10")
        frame_sub.pack(fill="x", expand=False) 

        # 버튼 하나로 통합
        btn_calc_all = ttk.Button(frame_sub, text="📊 소요량 산출", command=self.open_unified_calc_popup)
        btn_calc_all.pack(fill="x", ipady=8)

        # -------------------------------------------------------
        # [우측 패널] 시뮬레이션 설정 및 리스트
        # -------------------------------------------------------
        frame_sim = ttk.LabelFrame(right_panel, text=" 2. 사전투표소별 설정 및 현황 ", padding="10")
        frame_sim.pack(fill="both", expand=True)
        
        # [신규] 사전투표율 시뮬레이션 대시보드
        self.create_dashboard_ui(frame_sim)

        # 팁 문구
        lbl_tip = ttk.Label(frame_sim, text="💡 목록을 더블 클릭하여 장비 수나 조정률을 개별 수정할 수 있습니다.", 
                            foreground="gray", font=("맑은 고딕", 9))
        lbl_tip.pack(anchor="w", pady=(0, 5))

        # 트리뷰 (리스트) 영역
        tree_frame = ttk.Frame(frame_sim)
        tree_frame.pack(fill="both", expand=True)
        
        columns = ("station", "elect_diff", "intra", "extra", "rate_merged")
        # 줄무늬 스타일 적용
        style.configure("Treeview", rowheight=25)
        style.map("Treeview", background=[('selected', '#3B5BDB')])
        
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
        self.tree.tag_configure('even', background='#F0F4F8')
        self.tree.tag_configure('odd', background='white')
        
        self.tree.heading("station", text="사전투표소명")
        self.tree.heading("elect_diff", text="선거인수 변동")
        self.tree.heading("intra", text="관내장비")
        self.tree.heading("extra", text="관외장비")
        self.tree.heading("rate_merged", text="증가율(관내/외)") 
        
        self.tree.column("station", width=120)
        self.tree.column("elect_diff", width=100, anchor="center")
        self.tree.column("intra", width=80, anchor="center")
        self.tree.column("extra", width=80, anchor="center")
        self.tree.column("rate_merged", width=150, anchor="center")
        
        scrollbar_tree = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar_tree.set)

        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar_tree.pack(side="right", fill="y")
        
        self.tree.bind("<Double-1>", self.on_tree_double_click)
        
        # 상태 표시줄
        self.lbl_status = ttk.Label(self.root, text=" 준비됨", relief="sunken", anchor="w", font=("맑은 고딕", 9))
        self.lbl_status.pack(side="bottom", fill="x")

    def reset_all(self):
        # 1. 사용자 확인
        if not messagebox.askyesno("초기화 확인", "업로드한 파일과 목록을 모두 초기화하시겠습니까?\n(작업 중인 내용은 사라집니다.)"):
            return

        # 2. 내부 데이터 변수 초기화
        self.vote_files = []
        self.cached_data = {} 
        self.equipment_file = None
        self.file_past_elect = None   
        self.file_recent_elect = None 
        self.station_data = {}
        self.region_name = ""
        
        # 3. UI 텍스트 초기화
        self.lbl_file_count.config(text="파일 없음", foreground="gray")
        self.lbl_equip_status.config(text="파일 미선택 (기본값: 1대 적용)", foreground="gray")
        self.lbl_elect_status.config(text="파일 미선택 (변동률 미적용)", foreground="gray")
        
        # 4. 트리뷰(리스트) 비우기
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        # [추가] 시뮬레이션 관련 집계 변수 초기화
        self.total_past_voters = 0
        self.total_past_intra = 0
        self.total_past_extra = 0
        self.total_recent_electors = 0
        self.total_past_electors = 0
        self.past_turnout_rate = 0.0

        # [추가] 대시보드 UI(라벨 및 입력창) 텍스트 초기화
        self.lbl_past_info.config(text="직전 사전투표율: - % (총 -명 / 관내 -명 / 관외 -명)")
        self.entry_predict_rate.delete(0, tk.END)
        self.lbl_predict_details.config(text="% (총 -명 / 관내 -명 / 관외 -명)")

        # 5. 슬라이더 및 증감률 입력창 초기화
        self.reset_rate_zero()

        # 6. 로그 남기기
        self.log("=== 모든 데이터가 초기화되었습니다 ===")
        messagebox.showinfo("완료", "초기화되었습니다.")

    def select_past_file(self):
        file = filedialog.askopenfilename(title="과거 선거인수 파일 (A열:동명, B열:인수)", filetypes=[("Excel Files", "*.xlsx *.xls")])
        if file:
            self.file_past_elect = file
            self._update_elect_status()

    def select_recent_file(self):
        file = filedialog.askopenfilename(title="최근 선거인수 파일 (A열:동명, B열:인수)", filetypes=[("Excel Files", "*.xlsx *.xls")])
        if file:
            self.file_recent_elect = file
            self._update_elect_status()

    def _update_elect_status(self):
        # 두 파일 상태 확인 및 스캔 트리거
        p = "✅" if self.file_past_elect else "❌"
        r = "✅" if self.file_recent_elect else "❌"
        
        self.lbl_elect_status.config(text=f"과거: {p} / 최근: {r}", foreground="blue" if (p=="✅" and r=="✅") else "red")
        
        if self.file_past_elect and self.file_recent_elect:
            self.log("두 선거인수 파일 준비됨. 변동률 계산 시작...")
            self.scan_stations()

    # [수정된 코드] 스레드 안전(Thread-Safe) 로그 함수
    def log(self, msg):
        # 1. 콘솔 출력은 스레드와 상관없으므로 즉시 실행
        print(f"[Log] {msg}")
        
        # 2. UI 업데이트는 메인 스레드(root)가 처리하도록 큐(after)에 등록
        # 0ms 후에 _update_status_ui 함수를 실행하라는 명령
        self.root.after(0, self._update_status_ui, msg)

    # [추가된 코드] 실제 UI를 변경하는 내부 함수 (메인 스레드에서만 실행됨)
    def _update_status_ui(self, msg):
        if hasattr(self, 'lbl_status'):
            self.lbl_status.config(text=f" 📢 {msg}")
            # update_idletasks()는 제거 (after가 이벤트 루프를 타므로 불필요)

    def select_vote_files(self):
        files = filedialog.askopenfilenames(title="투표 데이터 선택", filetypes=[("Excel/CSV Files", "*.xlsx *.xls *.csv")])
        if files:
            self.vote_files = files
            self.cached_data = {} # [최적화] 새 파일 선택 시 캐시 초기화
            self.lbl_file_count.config(text=f"✅ {len(files)}개 파일 로드됨", foreground="blue")
            self.log(f"{len(files)}개 파일 선택됨. 데이터 로드 및 스캔 시작...")
            self.scan_stations()

    def select_equip_file(self):
        file = filedialog.askopenfilename(title="장비현황 파일 선택", filetypes=[("Excel Files", "*.xlsx *.xls")])
        if file:
            self.equipment_file = file
            self.lbl_equip_status.config(text=f"✅ {os.path.basename(file)}", foreground="blue")
            self.log(f"장비 파일 로드됨. 목록 업데이트 중...")
            self.scan_stations() 

    def _ensure_data_loaded(self):
        # [최적화] 파일이 캐시에 없으면 읽어서 저장
        for file in self.vote_files:
            if file in self.cached_data:
                continue
                
            try:
                day, time, header_row = self.get_file_info_header(file)
                if day is None: continue

                if file.endswith('.csv'):
                    try: df = pd.read_csv(file, header=header_row, encoding='cp949')
                    except: df = pd.read_csv(file, header=header_row, encoding='utf-8')
                else:
                    df = pd.read_excel(file, header=header_row)

                if '사전투표소명' in df.columns:
                    df = df.dropna(subset=['사전투표소명'])
                    # 공통 전처리: 합계/소계 제거
                    if '읍면동명' in df.columns:
                        temp_col = df['읍면동명'].astype(str).str.replace(' ', '')
                        mask = temp_col.str.contains('합계|소계|총계|누계', na=False)
                        df = df[~mask].copy()
                    
                    # 숫자 변환 미리 수행
                    for col in ['관내사전투표자수', '관외사전투표자수']:
                        if col in df.columns and df[col].dtype == 'object':
                            df[col] = df[col].astype(str).str.replace(',', '').str.strip()
                            df[col] = pd.to_numeric(df[col], errors='coerce')

                    self.cached_data[file] = (df, day, time)
            except Exception as e:
                self.log(f"파일 로드 실패({os.path.basename(file)}): {e}")

    def scan_stations(self):
        if not self.vote_files:
            return

        self._ensure_data_loaded() # [최적화] 데이터 로드 보장

        station_list = []  
        seen = set()

        station_past_data = {}       
        
        # 캐시된 데이터에서 투표소 추출
        for file in self.vote_files:
            if file not in self.cached_data: continue
            
            df, _, _ = self.cached_data[file]
            stations = df['사전투표소명'].unique()
            
            for s in stations:
                s_str = str(s).strip()
                if s_str and s_str != 'nan':
                    if s_str not in seen:
                        seen.add(s_str)
                        station_list.append(s_str)

        # [수정] 사용자 지정 서식(C열:이름, F열:관내, G열:관외) 맞춤 로직
        equip_map = {}
        if self.equipment_file:
            try:
                # 1. 파일 읽기 (헤더 없이 읽음)
                if self.equipment_file.endswith('.csv'):
                    try: df_raw = pd.read_csv(self.equipment_file, header=None, encoding='cp949')
                    except: df_raw = pd.read_csv(self.equipment_file, header=None, encoding='utf-8')
                else:
                    df_raw = pd.read_excel(self.equipment_file, header=None)

                # [추가] 3행(Index 2)에서 지역 이름 추출 로직
                try:
                    # 3행의 데이터 중 '값'이 있는 첫 번째 칸을 지역 이름으로 간주
                    row_3_vals = df_raw.iloc[2].astype(str).values
                    for v in row_3_vals:
                        v_clean = v.strip().replace('nan', '')
                        if v_clean:
                            self.region_name = v_clean
                            break
                except:
                    self.region_name = ""

                # 2. 데이터 시작 행 찾기 (C열에 '읍면동'이나 '투표소'가 나오는 줄)
                start_row_idx = 0
                for idx, row in df_raw.head(15).iterrows():
                    # 엑셀 C열은 인덱스 2
                    c_col_val = str(row[2]).replace(" ", "")
                    if "읍면동" in c_col_val or "투표소" in c_col_val:
                        start_row_idx = idx + 1 # 헤더 다음 줄부터 데이터
                        break
                
                # 3. 데이터 추출 (C열=2, F열=5, G열=6)
                for idx in range(start_row_idx, len(df_raw)):
                    row = df_raw.iloc[idx]
                    
                    # C열: 투표소명
                    st_name = str(row[2]).strip()
                    if st_name == 'nan' or not st_name: continue
                    if '합계' in st_name or '소계' in st_name: continue

                    # 숫자 정제 함수
                    def parse_count(val):
                        try:
                            txt = str(val).split('(')[0].replace(',', '').replace('대', '').strip()
                            return int(float(txt))
                        except:
                            return 1 

                    # F열(5): 관내 장비수, G열(6): 관외 장비수
                    intra_count = parse_count(row[5])
                    extra_count = parse_count(row[6])

                    equip_map[st_name] = {'intra': intra_count, 'extra': extra_count}

                self.log(f"장비 파일 로드 완료: {len(equip_map)}개소 (C,F,G열 기준)")

            except Exception as e:
                self.log(f"장비 파일 읽기 오류: {e}")
        
        # [수정] 선거인수 변동률 계산 로직 (두 파일 병합)
        electorate_rates = {}
        electorate_diffs = {}
        
        if self.file_past_elect and self.file_recent_elect:
            try:
                # 데이터 추출 내부 함수 (A열: 동명, B열: 숫자 라고 가정)
                # [수정] 데이터 추출 내부 함수 (A열: 읍면동명, D열: 선거인수)
                def load_elect_data(path):
                    data_map = {}
                    try:
                        # 1. 헤더 없이 읽어서 데이터 위치 찾기
                        if path.endswith('.csv'):
                            try: df = pd.read_csv(path, header=None, encoding='cp949')
                            except: df = pd.read_csv(path, header=None, encoding='utf-8')
                        else:
                            df = pd.read_excel(path, header=None)
                        
                        start_row = 0
                        # 2. '읍면동명'이 있는 행 찾기 (헤더 위치 검색)
                        for idx, row in df.head(15).iterrows():
                            # A열(0번 인덱스) 확인
                            if "읍면동명" in str(row[0]):
                                start_row = idx + 1
                                break
                        
                        # 3. 데이터 추출
                        for idx in range(start_row, len(df)):
                            row = df.iloc[idx]
                            
                            # A열(0): 동 이름
                            k_raw = str(row[0])
                            if pd.isna(row[0]) or k_raw.strip() == '' or k_raw == 'nan': continue
                            
                            k = k_raw.strip().replace(" ", "")
                            if '합계' in k or '소계' in k: continue # 합계 행 제외
                            
                            # D열(3): 선거인수 (예: "21,412\n(25, 12)")
                            v_raw = str(row[3])
                            
                            try:
                                # 줄바꿈(\n)이나 괄호(() 앞부분의 숫자만 가져오기
                                v_str = v_raw.split('\n')[0].split('(')[0]
                                v_str = v_str.replace(',', '').strip()
                                v = float(v_str)
                                
                                if v > 0:
                                    data_map[k] = v
                            except:
                                continue
                                
                    except Exception as e:
                        print(f"선거인수 파일 읽기 실패({path}): {e}")
                        
                    return data_map

                past_map = load_elect_data(self.file_past_elect)
                recent_map = load_elect_data(self.file_recent_elect)
                
                # 두 맵을 비교하여 증감률 및 차이 계산
                count_matched = 0
                for dong_name, recent_val in recent_map.items():
                    if dong_name in past_map:
                        past_val = past_map[dong_name]
                        if past_val > 0:
                            # 1. 시뮬레이션용 비율 계산 (유지)
                            rate_val = ((recent_val - past_val) / past_val) * 100
                            electorate_rates[dong_name] = rate_val
                            
                            # 2. [추가] 화면 표시용 인원 차이 계산
                            diff_val = int(recent_val - past_val)
                            electorate_diffs[dong_name] = diff_val
                            
                            count_matched += 1
                # [추가] 전체 선거인수 및 투표율 집계 로직
                self.total_recent_electors = sum(recent_map.values()) # 이번 선거인수 총합
                self.total_past_electors = sum(past_map.values())     # [수정] 과거 선거인수 총합 (멤버변수 저장)
                
                # 과거 투표자 수 집계 (캐시된 데이터 활용) 부분 찾아서 아래 코드로 교체
                
                temp_voter_sum = 0
                temp_intra_sum = 0 # [신규]
                temp_extra_sum = 0 # [신규]
                
                if self.vote_files:
                    try:
                        all_dfs = []
                        for file in self.vote_files:
                            if file in self.cached_data:
                                df_t, d_t, t_t = self.cached_data[file]
                                temp = df_t.copy()
                                temp['Day'] = d_t
                                all_dfs.append(temp)
                        
                        if all_dfs:
                            full_df = pd.concat(all_dfs)
                            # (투표소, 일차)별 최대값(누적)을 찾아서 합산
                            grp = full_df.groupby(['사전투표소명', 'Day'])[['관내사전투표자수', '관외사전투표자수']].max()
                            
                            # [수정] 관내/관외 전체 합계
                            self.total_past_intra = grp['관내사전투표자수'].sum()
                            self.total_past_extra = grp['관외사전투표자수'].sum()
                            temp_voter_sum = self.total_past_intra + self.total_past_extra
                            
                            # [신규] 개별 투표소 과거 데이터 저장 (역산용)
                            station_past_data = {}
                            grp_flat = grp.groupby('사전투표소명')[['관내사전투표자수', '관외사전투표자수']].sum()
                            
                            for st_name, row in grp_flat.iterrows():
                                station_past_data[st_name] = {
                                    'past_intra': row['관내사전투표자수'],
                                    'past_extra': row['관외사전투표자수']
                                }
                            
                            # [수정] 관내/관외 각각 합계 구하기
                            temp_intra_sum = grp['관내사전투표자수'].sum()
                            temp_extra_sum = grp['관외사전투표자수'].sum()
                            temp_voter_sum = temp_intra_sum + temp_extra_sum
                            
                    except Exception as e:
                        print(f"투표자 집계 오류: {e}")

                self.total_past_voters = temp_voter_sum
                self.total_past_intra = temp_intra_sum # [신규] 저장
                self.total_past_extra = temp_extra_sum # [신규] 저장
                
                if self.total_past_electors > 0:
                    self.past_turnout_rate = (self.total_past_voters / self.total_past_electors) * 100
                else:
                    self.past_turnout_rate = 0.0
                    

                self.log(f"변동률 계산 완료: {count_matched}개 동 매칭됨")
                
            except Exception as e:
                self.log(f"선거인수 파일 처리 오류: {e}")

        # =================================================================
        # [수정] 지역 불일치 감지 및 자동 초기화 (인구 + 장비 모두 포함)
        # =================================================================
        # 감지 조건: 인구 데이터가 있거나, 장비 지역명(region_name)이 있을 때 검사
        check_needed = False
        is_mismatch = False
        
        # 1. 인구 데이터와 비교
        if electorate_rates and station_list:
            check_needed = True
            is_pop_match = False
            for st in station_list:
                st_clean = st.replace(" ", "")
                for dong in electorate_rates.keys():
                    if dong in st_clean:
                        is_pop_match = True
                        break
                if is_pop_match: break
            if not is_pop_match: is_mismatch = True

        # 2. 장비 데이터(지역명)와 비교 (장비 파일에 지역명이 감지된 경우)
        if not is_mismatch and self.region_name and station_list:
            check_needed = True
            is_equip_match = False
            for st in station_list:
                # 투표소 이름에 장비파일 지역명(예: "유성구")이 포함되는지 체크는 어려우므로
                # 보통 동 이름 매칭 실패 시 함께 처리되지만, 안전을 위해 로직은 남겨둠
                pass 
            # (장비 파일만으로는 동 이름을 알 수 없으므로, 위 인구 데이터 불일치 시 함께 날리는 방식으로 처리)

        # 결론: 불일치 발생 시 초기화 수행
        if check_needed and is_mismatch:
            messagebox.showwarning("지역 데이터 불일치", 
                "새로 로드한 [투표 데이터]가 기존 [인구/장비 데이터]와 지역이 다릅니다.\n\n"
                "오류 방지를 위해 기존에 등록된\n"
                "1. 운용장비 현황 파일\n"
                "2. 인구수 통보 파일\n"
                "을 모두 자동으로 해제합니다.\n\n"
                "해당 지역에 맞는 파일들을 다시 업로드해주세요.")
            
            # --- [1] 인구 데이터 초기화 ---
            self.file_past_elect = None
            self.file_recent_elect = None
            electorate_rates = {} 
            electorate_diffs = {}
            self._update_elect_status() # UI 갱신 (X 표시)
            
            # --- [2] 장비 데이터 초기화 (추가된 부분) ---
            self.equipment_file = None
            self.region_name = ""
            # UI 갱신 (파일 미선택 상태로 복구)
            self.lbl_equip_status.config(text="파일 미선택 (기본값: 1대 적용)", foreground="gray")
            
            self.log("지역 불일치로 [인구] 및 [장비] 데이터 자동 초기화됨")
        # ================================================================= 

        for item in self.tree.get_children():
            self.tree.delete(item)
        
        sorted_stations = station_list
        self.station_data = {} 
        # [수정] 소수점 1자리까지 정확히 가져오도록 변경
        current_global_rate = round(float(self.var_rate.get()), 1)

        # [수정] enumerate를 사용하여 인덱스(i)를 함께 가져옴
        for i, st in enumerate(sorted_stations):
            # 1. 장비 매칭
            matched_data = None
            if st in equip_map:
                matched_data = equip_map[st]
            else:
                for k, v in equip_map.items():
                    if str(k) in st or st in str(k): 
                        matched_data = v
                        break
            
            if matched_data:
                intra = matched_data['intra']
                extra = matched_data['extra']
            else:
                intra = 1
                extra = 1
            
            # 2. 선거인수 증감률 매칭
            elect_display = "-"
            elect_rate = 0 
            
            if electorate_rates:
                st_clean = st.replace(" ", "")
                for dong_name, e_rate in electorate_rates.items():
                    if dong_name in st_clean:
                        # [수정] 제2, 제3... 등 '제2' 이상의 투표소는 인구 변동 미적용 (본소에만 적용)
                        # 정규식: '제' 뒤에 2~9 숫자가 오고 뒤이어 '사전'이 붙는 패턴 찾기 (예: 제2사전, 제3사전)
                        if re.search(r'제[2-9]사전', st_clean):
                            elect_rate = 0
                            elect_display = "-" # 표기도 제외
                        else:
                            # 제1이거나 숫자가 없는 경우만 적용
                            elect_rate = e_rate
                            diff = electorate_diffs.get(dong_name, 0)
                            
                            if diff > 0: elect_display = f"+ {diff:,}" 
                            elif diff < 0: elect_display = f"- {abs(diff):,}"
                            else: elect_display = "-" 
                        break
            
            # [신규] 과거 투표자 수 매칭 (없으면 0)
            p_intra = 0
            p_extra = 0
            if 'station_past_data' in locals() and st in station_past_data:
                p_intra = station_past_data[st]['past_intra']
                p_extra = station_past_data[st]['past_extra']

            # 데이터 저장
            self.station_data[st] = {
                'intra': intra, 'extra': extra, 
                'rate_intra': current_global_rate,
                'rate_extra': current_global_rate,
                'elect_rate': elect_rate,
                'org_intra': intra, 'org_extra': extra,
                
                # [추가됨] 개별 역산을 위한 과거 데이터
                'past_intra': p_intra, 
                'past_extra': p_extra
            }
            
            rate_txt = self._get_merged_rate_text(current_global_rate, current_global_rate)

            # [핵심 수정] 짝수(0,2,4...)는 'even', 홀수(1,3,5...)는 'odd' 태그 적용
            row_tag = 'even' if i % 2 == 0 else 'odd'
            # [변경] 화면 표시용 이름 생성 ('사전투표소' 제거)
            st_disp = st.replace("사전투표소", "")
            self.tree.insert("", "end", iid=st, values=(st_disp, elect_display, intra, extra, rate_txt), tags=(row_tag,))
        self._update_dashboard_info()    
        self.log(f"목록 갱신 완료: 총 {len(sorted_stations)}개 투표소")

    def on_tree_double_click(self, event):
        try:
            # 1. 클릭한 위치(행/열) 파악
            region = self.tree.identify("region", event.x, event.y)
            if region != "cell": return 
            
            item_id = self.tree.identify_row(event.y)
            column = self.tree.identify_column(event.x)
            
            if not item_id: return
            
            # [수정 1] item_values는 '선거인수(elect_disp)'를 가져오기 위해 꼭 필요하므로 주석 해제
            item_values = self.tree.item(item_id)['values']
            if not item_values: return
            
            # [수정 2] 이름은 '화면에 보이는 값(values[0])'이 아닌 '고유 ID(item_id, 풀네임)' 사용
            st_name = item_id 
            
            # 안전장치
            if st_name not in self.station_data:
                return

            # 3. 데이터 가져오기
            data = self.station_data[st_name]
            curr_intra = data['intra']
            curr_extra = data['extra']
            org_intra = data['org_intra']
            org_extra = data['org_extra']
            
            val_rate_intra = data['rate_intra']
            val_rate_extra = data['rate_extra']
            
            elect_disp = item_values[1] # 선거인수 표기 유지

            # 화면 표시용 텍스트 생성 내부함수
            def get_display_text(val, org_val):
                return f"{org_val} → {val}" if val != org_val else str(val)

            # 4. 컬럼별 수정 로직
            if column == '#3': # 관내 장비
                new_intra = simpledialog.askinteger("관내 장비 수정", f"[{st_name}]\n관내 장비 수:", 
                                                  initialvalue=curr_intra, minvalue=1, maxvalue=50, parent=self.root)
                if new_intra is not None:
                    self.station_data[st_name]['intra'] = new_intra
                    disp_intra = get_display_text(new_intra, org_intra)
                    disp_extra = get_display_text(curr_extra, org_extra)
                    st_disp = st_name.replace("사전투표소", "")
                    
                    # [수정 3] 텍스트 형식으로 변환하여 5번째 컬럼에 적용
                    rate_txt = self._get_merged_rate_text(val_rate_intra, val_rate_extra)
                    self.tree.item(item_id, values=(st_disp, elect_disp, disp_intra, disp_extra, rate_txt))
                    
                    self.log(f"{st_name} 관내 장비 변경: {new_intra}대")
                    
            elif column == '#4': # 관외 장비
                new_extra = simpledialog.askinteger("관외 장비 수정", f"[{st_name}]\n관외 장비 수:", 
                                                  initialvalue=curr_extra, minvalue=1, maxvalue=50, parent=self.root)
                if new_extra is not None:
                    self.station_data[st_name]['extra'] = new_extra
                    disp_intra = get_display_text(curr_intra, org_intra)
                    disp_extra = get_display_text(new_extra, org_extra)
                    st_disp = st_name.replace("사전투표소", "")
                    
                    # [수정 3] 텍스트 형식으로 변환하여 5번째 컬럼에 적용
                    rate_txt = self._get_merged_rate_text(val_rate_intra, val_rate_extra)
                    self.tree.item(item_id, values=(st_disp, elect_disp, disp_intra, disp_extra, rate_txt))
                    
                    self.log(f"{st_name} 관외 장비 변경: {new_extra}대")
                    
            elif column == '#5': # 조정률(통합) 수정
                self._open_rate_input_dialog(st_name, item_id, elect_disp, curr_intra, curr_extra, org_intra, org_extra)
            
            else:
                messagebox.showinfo("알림", "수정 가능한 항목(장비 수, 조정률)을 더블 클릭해주세요.", parent=self.root)

        except Exception as e:
            print(f"더블 클릭 오류: {e}")
            import traceback
            traceback.print_exc()

    def get_file_info_header(self, file_path):
        try:
            if file_path.endswith('.csv'):
                try: df_meta = pd.read_csv(file_path, header=None, nrows=10, encoding='cp949')
                except: df_meta = pd.read_csv(file_path, header=None, nrows=10, encoding='utf-8')
            else:
                df_meta = pd.read_excel(file_path, header=None, nrows=10)
            
            day, time = None, None
            header_idx = 3

            for idx, row in df_meta.iterrows():
                row_str = " ".join(row.astype(str).values)
                if day is None:
                    match_day = re.search(r'\[(\d+)일차\]', row_str)
                    match_time = re.search(r'\[(\d{1,2}):(\d{2})\]', row_str)
                    if match_day: day = int(match_day.group(1))
                    if match_time: time = int(match_time.group(1))
                if "사전투표소명" in row_str or "읍면동명" in row_str:
                    header_idx = idx
            return day, time, header_idx
        except:
            return None, None, 3

    def run_simulation(self):
        if not self.vote_files:
            messagebox.showwarning("주의", "투표 데이터 파일이 없습니다.")
            return

        # 1. 로딩 팝업창 생성
        self.loading_win = tk.Toplevel(self.root)
        self.loading_win.title("처리 중")
        self.loading_win.geometry("300x100")
        self.loading_win.resizable(False, False)
        # 팝업이 떠 있는 동안 메인 창 조작 금지 (모달)
        self.loading_win.grab_set() 
        
        # 화면 중앙 배치
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - 150
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - 50
        self.loading_win.geometry(f"+{x}+{y}")

        lbl = ttk.Label(self.loading_win, text="데이터 분석 및 시각화 중입니다...\n잠시만 기다려 주세요.", justify="center")
        lbl.pack(pady=20)
        
        # 프로그레스바 (왔다갔다 하는 모드)
        pb = ttk.Progressbar(self.loading_win, mode='indeterminate')
        pb.pack(fill="x", padx=20, pady=(0, 20))
        pb.start(10)

        # 2. 별도 스레드에서 무거운 작업 실행
        # daemon=True로 설정하여 메인 프로그램 종료 시 같이 종료되게 함
        t = threading.Thread(target=self._execute_simulation, daemon=True)
        t.start()

    def _execute_simulation(self):
        try:
            # [기존 설정 유지]
            import matplotlib
            matplotlib.use('Agg')
            import warnings
            warnings.simplefilter(action='ignore', category=FutureWarning)

            label = "통합 분석"
            self.log(f"시뮬레이션 시작: {label}")
            
            self._ensure_data_loaded() 
            
            all_data = []
            
            # [기존 데이터 로드 및 계산 로직 유지]
            for file in self.vote_files:
                if file not in self.cached_data: continue
                
                try:
                    org_df, day, time = self.cached_data[file]
                    df = org_df.copy() 
                    
                    df['사전투표소명'] = df['사전투표소명'].astype(str).str.strip()
                    
                    user_rate_intra_map = {name: data.get('rate_intra', 0) for name, data in self.station_data.items()}
                    user_rate_extra_map = {name: data.get('rate_extra', 0) for name, data in self.station_data.items()}
                    
                    user_rates_intra = df['사전투표소명'].map(user_rate_intra_map).fillna(0)
                    user_rates_extra = df['사전투표소명'].map(user_rate_extra_map).fillna(0)
                    
                    elect_rate_map = {name: data.get('elect_rate', 0) for name, data in self.station_data.items()}
                    elect_rates = df['사전투표소명'].map(elect_rate_map).fillna(0)
                    
                    factor_intra = (1 + (elect_rates / 100.0)) * (1 + (user_rates_intra / 100.0))
                    factor_extra = (1 + (user_rates_extra / 100.0))
                    
                    df['관내사전투표자수'] = df['관내사전투표자수'] * factor_intra
                    df['관외사전투표자수'] = df['관외사전투표자수'] * factor_extra
                            
                    df['일차'] = day
                    df['시간대'] = time
                    all_data.append(df)
                except Exception as e:
                    self.log(f"데이터 처리 오류({os.path.basename(file)}): {e}")

            if not all_data:
                self.root.after(0, lambda: messagebox.showerror("오류", "유효한 데이터가 없습니다."))
                self.root.after(0, self.loading_win.destroy)
                return

            final_df = pd.concat(all_data, ignore_index=True)

            original_order = []
            seen = set()
            for temp_df in all_data:
                stats = temp_df['사전투표소명'].unique()
                for s in stats:
                    if s not in seen:
                        seen.add(s)
                        original_order.append(s)
            
            final_df['사전투표소명'] = pd.Categorical(
                final_df['사전투표소명'], categories=original_order, ordered=True
            )
            
            duplicates = final_df[final_df.duplicated(subset=['사전투표소명', '일차', '시간대'], keep=False)]
            if not duplicates.empty:
                final_df = final_df.drop_duplicates(subset=['사전투표소명', '일차', '시간대'])

            final_df = final_df.sort_values(by=['사전투표소명', '일차', '시간대'])
            
            final_df['시간대별_관내투표자수'] = final_df.groupby(['사전투표소명', '일차'], observed=True)['관내사전투표자수'].diff()
            final_df['시간대별_관외투표자수'] = final_df.groupby(['사전투표소명', '일차'], observed=True)['관외사전투표자수'].diff()
            
            for (st, day), group in final_df.groupby(['사전투표소명', '일차'], observed=True):
                first_idx = group.index[0]
                final_df.loc[first_idx, '시간대별_관내투표자수'] = final_df.loc[first_idx, '관내사전투표자수']
                final_df.loc[first_idx, '시간대별_관외투표자수'] = final_df.loc[first_idx, '관외사전투표자수']

            def get_equip_info(row, type_):
                st = row['사전투표소명']
                if st in self.station_data:
                    return self.station_data[st][type_], self.station_data[st][f'org_{type_}']
                return 1, 1

            final_df[['관내장비수', '원본_관내장비수']] = final_df.apply(lambda x: pd.Series(get_equip_info(x, 'intra')), axis=1)
            final_df[['관외장비수', '원본_관외장비수']] = final_df.apply(lambda x: pd.Series(get_equip_info(x, 'extra')), axis=1)

            final_df['관내_혼잡도'] = final_df['시간대별_관내투표자수'] / final_df['관내장비수']
            final_df['관외_혼잡도'] = final_df['시간대별_관외투표자수'] / final_df['관외장비수']

            final_df = final_df.loc[:, ~final_df.columns.str.contains('^Unnamed')]
            
            # [짧은 이름 생성] 시각화 및 엑셀 저장 시 사용
            final_df['short_name'] = final_df['사전투표소명'].astype(str).str.replace('사전투표소', '').str.strip()

            timestamp = datetime.now().strftime('%Y%m%d_%H%M')
            
            # [수정] 실행 파일 경로 확인
            if getattr(sys, 'frozen', False):
                base_dir = os.path.dirname(os.path.abspath(sys.executable))
            else:
                base_dir = os.path.dirname(os.path.abspath(__file__))

            # [수정] '임시저장' 폴더 생성 로직 추가
            save_dir = os.path.join(base_dir, "임시저장")
            if not os.path.exists(save_dir):
                os.makedirs(save_dir)

            # 1. Raw 데이터 엑셀 저장 (경로를 save_dir로 변경)
            excel_name = f"시뮬레이션_결과_{timestamp}.xlsx"
            full_excel_path = os.path.join(save_dir, excel_name)
            final_df.to_excel(full_excel_path, index=False)
            self.log(f"Raw 데이터 저장 완료: {full_excel_path}")

            # =========================================================================
            # [핵심 수정] 2. '전체(평균)' 데이터 생성 로직을 여기로 이동!
            # =========================================================================
            try:
                # [수정] 모드에 따라 '전체' 데이터 집계 방식 분기 (투표자수=합계, 혼잡도=평균)
                grp = final_df.groupby(['사전투표소명', '시간대', 'short_name'], observed=True)
                mode_val = self.var_mode.get()
                
                # 집계할 컬럼들 구분
                cols_mean = ['관내_혼잡도', '관외_혼잡도', '관내장비수', '원본_관내장비수', '관외장비수', '원본_관외장비수']
                cols_sum = ['시간대별_관내투표자수', '시간대별_관외투표자수']
                
                if mode_val == "population":
                    # 투표자수 모드: 투표자수는 합계(sum), 나머지는 평균(mean)
                    agg_dict = {c: 'mean' for c in cols_mean}
                    agg_dict.update({c: 'sum' for c in cols_sum})
                    # agg 함수로 컬럼별 다른 연산 적용
                    df_mean = grp.agg(agg_dict).reset_index()
                else:
                    # 혼잡도 모드: 모두 평균(mean) (기존 방식)
                    all_cols = cols_mean + cols_sum
                    df_mean = grp[all_cols].mean().reset_index()
                    
                df_mean['일차'] = '전체'
                # 원본 final_df에 합치기
                final_df = pd.concat([final_df, df_mean], ignore_index=True)
                self.log("전체(평균) 데이터 계산 완료")
            except Exception as e:
                self.log(f"평균 데이터 생성 실패: {e}")

            # 3. 시각화 형태(히트맵) 엑셀 리포트 저장 (경로를 save_dir로 변경)
            report_name = f"시각화_리포트_{timestamp}.xlsx"
            full_report_path = os.path.join(save_dir, report_name)
            self.save_visual_excel(final_df, full_report_path)
            self.log(f"시각화 리포트 저장 완료: {full_report_path}")
            
            self.log("그래프 생성 중...")
            
            png_name = f"시뮬레이션_{timestamp}.png"
            full_png_path = os.path.join(save_dir, png_name)

            self.visualize_results(final_df, timestamp, full_png_path, mode='screen')
            
            def _finish():
                if hasattr(self, 'loading_win'): self.loading_win.destroy()
                messagebox.showinfo("완료", f"분석 완료!\n\n파일이 저장되었습니다:\n{full_png_path}")
                if platform.system() == 'Windows':
                    try: os.startfile(full_png_path)
                    except: pass
            
            self.root.after(0, _finish)

        except Exception as e:
            err_msg = str(e)
            import traceback
            traceback.print_exc()

            def _error():
                if hasattr(self, 'loading_win'): self.loading_win.destroy()
                self.log(f"치명적 오류: {err_msg}")
                messagebox.showerror("오류", f"작업 중 오류가 발생했습니다.\n{err_msg}")
                
            self.root.after(0, _error)

    def visualize_results(self, df, timestamp, save_name, mode='screen'):
        # 1. 폰트 설정
        system_name = platform.system()
        font_family = 'Malgun Gothic' if system_name == 'Windows' else 'AppleGothic'
        plt.rc('font', family=font_family)
        plt.rc('axes', unicode_minus=False)

        df['label_clean'] = df['short_name'] 
        
        # === [수정] 사용자가 선택한 모드 확인 ===
        # 라디오 버튼 변수값 확인
        mode_val = self.var_mode.get()
        
        if mode_val == "population":
            target_col_intra = '시간대별_관내투표자수'
            target_col_extra = '시간대별_관외투표자수'
            color_map = 'Oranges' # 주황색
            title_suffix = "(사전투표자 수)" # <--- 변경
        else:
            # 기본값: density
            target_col_intra = '관내_혼잡도'
            target_col_extra = '관외_혼잡도'
            color_map = 'Greens'  # 녹색 (기본)
            title_suffix = "(장비 1대당 발급자수)" # <--- 변경

        # 4. 시나리오 설정 (체크박스 값 반영 + 동적 컬럼)
        all_scenarios = [
            (1, '관내', 'label_clean', target_col_intra, '관내장비수', '원본_관내장비수', self.var_day1.get() and self.var_intra.get()),
            (1, '관외', 'label_clean', target_col_extra, '관외장비수', '원본_관외장비수', self.var_day1.get() and self.var_extra.get()),
            (2, '관내', 'label_clean', target_col_intra, '관내장비수', '원본_관내장비수', self.var_day2.get() and self.var_intra.get()),
            (2, '관외', 'label_clean', target_col_extra, '관외장비수', '원본_관외장비수', self.var_day2.get() and self.var_extra.get()),
            ('전체', '관내', 'label_clean', target_col_intra, '관내장비수', '원본_관내장비수', self.var_day_all.get() and self.var_intra.get()),
            ('전체', '관외', 'label_clean', target_col_extra, '관외장비수', '원본_관외장비수', self.var_day_all.get() and self.var_extra.get())
        ]
        
        # 활성화된 시나리오만 필터링
        active_scenarios = [s for s in all_scenarios if s[6]]
        if not active_scenarios: return

        unique_stations = df['사전투표소명'].unique()
        
        # [수정] _plot_page에 색상(cmap)과 제목접미사(title_suffix) 전달
        return self._plot_page(df, active_scenarios, unique_stations, filename=save_name, is_pdf=False, cmap=color_map, title_suffix=title_suffix)

    def save_visual_excel(self, df, filename):
        # === [추가] 모드 확인 및 설정 ===
        mode_val = self.var_mode.get()
        
        # [수정] 모드에 따라 엑셀 헤더 및 색상 설정
        if mode_val == "population":
            target_col_intra = '시간대별_관내투표자수'
            target_col_extra = '시간대별_관외투표자수'
            # 주황색 계열 (Excel Color Scale)
            start_c, mid_c, end_c = 'FFF5EB', 'FDAE6B', 'E6550D' 
            # 엑셀에 표시할 합계/평균 라벨 동적 설정
            total_label = '전체합계'
            row_stat_label = '시간대합계'
        else:
            target_col_intra = '관내_혼잡도'
            target_col_extra = '관외_혼잡도'
            # 녹색 계열 (기존)
            start_c, mid_c, end_c = 'F7FCF5', '74C476', '006D2C'
            total_label = '전체평균'
            row_stat_label = '시간대평균'

        # 1. 시나리오 정의 (동적 변수 적용)
        scenarios = [
            ('1일차_관내', 1, '관내', target_col_intra, '관내장비수', '원본_관내장비수', self.var_day1.get() and self.var_intra.get()),
            ('1일차_관외', 1, '관외', target_col_extra, '관외장비수', '원본_관외장비수', self.var_day1.get() and self.var_extra.get()),
            ('2일차_관내', 2, '관내', target_col_intra, '관내장비수', '원본_관내장비수', self.var_day2.get() and self.var_intra.get()),
            ('2일차_관외', 2, '관외', target_col_extra, '관외장비수', '원본_관외장비수', self.var_day2.get() and self.var_extra.get()),
            ('전체_관내', '전체', '관내', target_col_intra, '관내장비수', '원본_관내장비수', self.var_day_all.get() and self.var_intra.get()),
            ('전체_관외', '전체', '관외', target_col_extra, '관외장비수', '원본_관외장비수', self.var_day_all.get() and self.var_extra.get())
        ]

        # 2. 엑셀 작성 시작
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            for (sheet_name, day, type_name, value_col, eq_col, org_eq_col, active) in scenarios:
                if not active: continue

                # 데이터 필터링
                if str(day) == '전체':
                    df_day = df[df['일차'] == '전체']
                else:
                    df_day = df[df['일차'] == day]
                
                if df_day.empty: continue

                # 피벗 테이블 생성
                pivot = df_day.pivot_table(index=['short_name'], columns='시간대', values=value_col)
                
                # [수정] 모드에 따라 엑셀 통계(행/열) 방식 변경
                if mode_val == "population":
                    # 투표자수 모드: 가로/세로 모두 합계(Sum)
                    pivot[total_label] = pivot.sum(axis=1) 
                    avg_row = pivot.sum(axis=0)       
                else:
                    # 혼잡도 모드: 가로/세로 모두 평균(Mean)
                    pivot[total_label] = pivot.mean(axis=1)
                    avg_row = pivot.mean(axis=0)
                
                pivot.loc[row_stat_label] = avg_row
                
                # 컬럼 순서 정리: [전체합계/평균]을 맨 앞으로
                time_cols = sorted([c for c in pivot.columns if c != total_label])
                new_cols = [total_label] + time_cols
                pivot = pivot[new_cols]

                original_order = list(dict.fromkeys(df_day['short_name']))
                
                # pivot 테이블에 실제로 존재하는 투표소만 필터링
                station_rows = [name for name in original_order if name in pivot.index]
                
                new_rows = [row_stat_label] + station_rows
                pivot = pivot.reindex(new_rows)

                # 장비 정보 가져오기
                equip_data = df_day.drop_duplicates(subset=['short_name']).set_index('short_name')[[eq_col, org_eq_col]]
                
                # 엑셀용 데이터프레임 구성 (장비 컬럼 추가)
                final_sheet_df = pivot.copy()
                final_sheet_df.insert(0, '장비수', "") 

                for idx in final_sheet_df.index:
                    if idx == row_stat_label:
                        # [수정] 투표자수 모드일 때는 괄호(평균) 표시 제거
                        if mode_val == "population":
                            # 합계만 표시 (천단위 콤마는 엑셀 서식으로 처리되거나 여기서 문자열로 박아도 됨)
                            val = pivot.loc[idx, total_label]
                            final_sheet_df.loc[idx, total_label] = val # 숫자로 남겨둠 (나중에 서식 적용)
                        else:
                            # 혼잡도 모드는 기존처럼 (집중평균) 표시
                            target_hours = [c for c in pivot.columns if isinstance(c, (int, float)) and 11 <= c <= 18]
                            if target_hours:
                                mean_val = pivot.loc[idx, total_label]
                                focus_mean = pivot.loc[idx, target_hours].mean()
                                final_sheet_df.loc[idx, total_label] = f"{mean_val:.1f}\n({focus_mean:.1f})"
                            else:
                                final_sheet_df.loc[idx, total_label] = f"{pivot.loc[idx, total_label]:.1f}"
                                
                        final_sheet_df.loc[idx, '장비수'] = "합계" if mode_val == "population" else "평균"
                    else:
                        # 장비수 텍스트 생성
                        try:
                            curr = int(equip_data.loc[idx, eq_col])
                            org = int(equip_data.loc[idx, org_eq_col])
                            txt = f"{org} → {curr}" if curr != org else f"{curr}"
                            final_sheet_df.loc[idx, '장비수'] = txt
                        except:
                            final_sheet_df.loc[idx, '장비수'] = "-"
                        
                # 엑셀 시트에 쓰기
                final_sheet_df.to_excel(writer, sheet_name=sheet_name)
                
                # --- 스타일링 (openpyxl) ---
                ws = writer.sheets[sheet_name]
                
                font_basic = Font(name='맑은 고딕', size=10)
                align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
                border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                border_thick_blue = Border(left=Side(style='medium', color='0000FF'), right=Side(style='medium', color='0000FF'), 
                                           top=Side(style='medium', color='0000FF'), bottom=Side(style='medium', color='0000FF'))

                max_row = ws.max_row
                max_col = ws.max_column
                
                for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                    for cell in row:
                        cell.font = font_basic
                        cell.alignment = align_center
                        cell.border = border_thin
                        # [추가] 투표자수 모드일 때 숫자 셀에 천단위 콤마 서식 적용
                        if mode_val == "population" and isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0'

                # 2. 헤더 스타일 (1행)
                for cell in ws[1]:
                    cell.font = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

                # 3. 인덱스 열 스타일 (A열: 투표소명, B열: 장비수)
                for row in range(2, max_row + 1):
                    ws.cell(row=row, column=1).font = Font(name='맑은 고딕', size=10, bold=True) 
                    ws.cell(row=row, column=2).font = Font(name='맑은 고딕', size=9) 

                # 4. 조건부 서식 (히트맵)
                if mode_val == "population":
                    # [유지] 투표자 수 모드는 기존대로 상대값(Min-Max) 기준
                    rule = ColorScaleRule(start_type='min', start_color=start_c,
                                          mid_type='percentile', mid_value=50, mid_color=mid_c,
                                          end_type='max', end_color=end_c)
                else:
                    # [수정] 혼잡도(장비당) 모드는 절대값 100 기준 고정
                    rule = ColorScaleRule(start_type='num', start_value=0, start_color=start_c,
                                          mid_type='num', mid_value=50, mid_color=mid_c,
                                          end_type='num', end_value=100, end_color=end_c)
                
                range_string = f"{get_column_letter(3)}3:{get_column_letter(max_col)}{max_row}"
                ws.conditional_formatting.add(range_string, rule)

                # 5. 파란색 테두리 강조
                # 시간대 합계/평균 행 (2행)
                for col in range(1, max_col + 1):
                    ws.cell(row=2, column=col).border = Border(top=Side(style='medium', color='0000FF'), 
                                                               bottom=Side(style='medium', color='0000FF'),
                                                               left=Side(style='thin'), right=Side(style='thin'))
                    ws.cell(row=2, column=col).font = Font(name='맑은 고딕', bold=True)
                    
                    if col >= 3:
                        # [수정] 모드에 따라 2행(통계행)의 숫자 서식 다르게
                        if mode_val == "population":
                            ws.cell(row=2, column=col).number_format = '#,##0'
                        else:
                            ws.cell(row=2, column=col).number_format = '0.0'

                # 전체 합계/평균 열 (C열)
                for row in range(1, max_row + 1):
                    cell = ws.cell(row=row, column=3)
                    prev_border = cell.border
                    cell.border = Border(left=Side(style='medium', color='0000FF'), 
                                         right=Side(style='medium', color='0000FF'),
                                         top=prev_border.top, bottom=prev_border.bottom)

                ws.cell(row=2, column=3).border = border_thick_blue
                
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 10
                ws.column_dimensions['C'].width = 12
                for col in range(4, max_col + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 6
    
    def _read_equip_summary(self):
        """
        장비현황 파일의 D7(총 장비수), H7(예비수) 셀을 읽어옵니다.
        엑셀/CSV 모두 헤더 없이 읽어서 좌표로 접근합니다.
        D7 -> Row 6, Col 3
        H7 -> Row 6, Col 7
        """
        if not self.equipment_file:
            return None, None
            
        try:
            # 헤더 없이 읽어서 절대 좌표(행/열)로 접근
            if self.equipment_file.endswith('.csv'):
                try: df = pd.read_csv(self.equipment_file, header=None, encoding='cp949')
                except: df = pd.read_csv(self.equipment_file, header=None, encoding='utf-8')
            else:
                df = pd.read_excel(self.equipment_file, header=None)
            
            # 파일 크기가 D7, H7을 읽을 수 있는지 확인 (행 7개 이상, 열 8개 이상)
            if df.shape[0] < 7 or df.shape[1] < 8:
                return None, None
                
            # D7 (Index: [6, 3]) -> 총 장비수
            raw_total = str(df.iloc[6, 3])
            # H7 (Index: [6, 7]) -> 예비수
            raw_reserve = str(df.iloc[6, 7])
            
            def _clean_num(val):
                # 숫자 외 문자 제거 (콤마, '대' 등)
                import re
                txt = re.sub(r'[^0-9]', '', val)
                return int(txt) if txt else 0
                
            total = _clean_num(raw_total)
            reserve = _clean_num(raw_reserve)
            
            return total, reserve
            
        except Exception as e:
            print(f"장비 요약 정보 로드 실패: {e}")
            return None, None
        
    def open_balance_popup(self):
        if not self.vote_files:
            messagebox.showwarning("주의", "먼저 투표 데이터 파일을 로드해주세요.")
            return
            
        # [수정] 기본값 설정 로직
        # 1순위: 장비 파일의 D7(총보유), H7(예비) 값 사용
        # 2순위: 파일 없으면 기존 로직(화면 합계 + 5) 사용
        
        file_total, file_reserve = self._read_equip_summary()
        
        if file_total is not None and file_total > 0:
            default_total_assets = file_total
            default_reserve = file_reserve if file_reserve is not None else 5
            self.last_reserve_count = default_reserve 
        else:
            # 파일이 없거나 읽기 실패 시
            curr_allocated = sum([item['intra'] + item['extra'] for item in self.station_data.values()])
            default_total_assets = curr_allocated + self.last_reserve_count
            default_reserve = self.last_reserve_count
        
        # 팝업창 생성
        pop = tk.Toplevel(self.root)
        pop.title("장비 자동 배분 (통합 모드)")
        pop.geometry("350x260") # [변경] 메시지 삭제로 높이를 300 -> 260으로 줄임
        pop.resizable(False, False)
        
        # 화면 중앙 배치
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - 175
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - 130
        pop.geometry(f"+{x}+{y}")
        
        # [변경] 안내 문구만 남기고 출처 메시지 삭제
        ttk.Label(pop, text="보유한 [전체 장비 수]를 입력하세요.\n관내/관외 구분 없이 혼잡도에 따라 통합 배분합니다.", 
                  justify="center", foreground="gray").pack(pady=(15, 10))
        
        frame_input = ttk.Frame(pop, padding="20")
        frame_input.pack(fill="both", expand=True)
        
        # 입력 필드 생성 함수
        def create_entry(parent, label, default_val):
            frame = ttk.Frame(parent)
            frame.pack(fill="x", pady=8)
            ttk.Label(frame, text=label, width=15, font=("맑은 고딕", 10, "bold")).pack(side="left")
            entry = ttk.Entry(frame, justify="right", font=("맑은 고딕", 10))
            entry.insert(0, str(default_val))
            entry.pack(side="right", expand=True, fill="x")
            return entry
            
        entry_total = create_entry(frame_input, "총 보유 장비:", default_total_assets)
        entry_reserve = create_entry(frame_input, "예비 장비:", default_reserve)
        
        def _run():
            try:
                total_assets = int(entry_total.get())
                total_reserve = int(entry_reserve.get())
                
                self.last_reserve_count = total_reserve
                
                available = total_assets - total_reserve
                min_req = len(self.station_data) * 2
                
                if available < min_req:
                    msg = f"장비가 부족합니다!\n\n투표소 수: {len(self.station_data)}개\n최소 필요 장비: {min_req}대 (관내1+관외1)\n현재 가용 장비: {available}대"
                    messagebox.showerror("배분 불가", msg)
                    return
                    
                self.run_auto_balance(total_assets, total_reserve)
                pop.destroy()
                
            except ValueError:
                messagebox.showerror("오류", "유효한 숫자를 입력해주세요.")

        ttk.Button(pop, text="최적 배분 실행", command=_run).pack(fill="x", padx=20, pady=20)

    def open_unified_calc_popup(self):
        if not self.vote_files:
            messagebox.showwarning("주의", "먼저 투표 데이터 파일을 로드해주세요.")
            return

        # 팝업창 생성
        pop = tk.Toplevel(self.root)
        pop.title("물품 소요량 통합 산출")
        pop.geometry("350x380")
        pop.resizable(False, False)
        
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - 175
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - 190
        pop.geometry(f"+{x}+{y}")

        # === 1. 기표대 설정 영역 ===
        frame_booth = ttk.LabelFrame(pop, text=" [기표대] 1인당 예상 기표 시간 (초) ", padding="15")
        frame_booth.pack(fill="x", padx=15, pady=(15, 10))

        def create_input(parent, label, default_val):
            f = ttk.Frame(parent)
            f.pack(fill="x", pady=5)
            ttk.Label(f, text=label, width=12, font=("맑은 고딕", 9)).pack(side="left")
            entry = ttk.Entry(f, justify="right", width=10)
            entry.insert(0, str(default_val))
            entry.pack(side="right")
            return entry

        entry_booth_intra = create_input(frame_booth, "① 관내 시간:", "")
        entry_booth_extra = create_input(frame_booth, "② 관외 시간:", "")

        # === 2. 롤 용지 설정 영역 ===
        frame_roll = ttk.LabelFrame(pop, text=" [용지] 1롤당 발급 가능 인원 (명) ", padding="15")
        frame_roll.pack(fill="x", padx=15, pady=5)

        entry_roll_intra = create_input(frame_roll, "① 관내 기준:", "")
        entry_roll_extra = create_input(frame_roll, "② 관외 기준:", "")

        # === 3. 실행 로직 ===
        def _run_calculation():
            # 1. 입력값 파싱
            def _get_val(entry):
                try:
                    val = int(entry.get())
                    return val if val > 0 else 0
                except:
                    return 0

            b_time_i = _get_val(entry_booth_intra)
            b_time_e = _get_val(entry_booth_extra)
            r_cap_i = _get_val(entry_roll_intra)
            r_cap_e = _get_val(entry_roll_extra)
            
            calc_booth = (b_time_i > 0 and b_time_e > 0)
            calc_roll = (r_cap_i > 0 and r_cap_e > 0)
            
            if not calc_booth and not calc_roll:
                messagebox.showerror("입력 오류", "기표대(시간) 또는 롤 용지(용량) 중\n적어도 하나의 세트는 올바르게 입력해야 합니다.", parent=pop)
                return

            self._ensure_data_loaded()
            
            # --- 데이터 준비 ---
            temp_data = {}
            all_keys = set()

            for file in self.vote_files:
                if file not in self.cached_data: continue
                df, day, time = self.cached_data[file]
                if time is None: continue
                
                time_key = (day, time)
                all_keys.add(time_key)
                
                for _, row in df.iterrows():
                    st_name = str(row['사전투표소명']).strip()
                    if st_name not in self.station_data: continue

                    if st_name not in temp_data: temp_data[st_name] = {}
                    
                    d = self.station_data[st_name]
                    factor_i = (1 + d.get('elect_rate',0)/100.0) * (1 + d['rate_intra']/100.0)
                    factor_e = (1 + d['rate_extra']/100.0)

                    try:
                        raw_i = row['관내사전투표자수']
                        raw_e = row['관외사전투표자수']
                        
                        if isinstance(raw_i, str): raw_i = float(raw_i.replace(',', ''))
                        if isinstance(raw_e, str): raw_e = float(raw_e.replace(',', ''))
                        
                        val_i = float(raw_i) * factor_i
                        val_e = float(raw_e) * factor_e
                        
                        temp_data[st_name][time_key] = {'intra': val_i, 'extra': val_e}
                    except Exception as e: 
                        pass

            # --- 결과 계산 ---
            main_order = []
            for item_id in self.tree.get_children():
                main_order.append(item_id)
            
            target_stations = [st for st in main_order if st in temp_data]
            
            if not target_stations:
                messagebox.showerror("오류", "계산할 데이터가 없습니다.\n투표소 이름 매칭을 확인해주세요.", parent=pop)
                return

            sorted_keys = sorted(list(all_keys))
            from collections import defaultdict
            day_groups = defaultdict(list)
            for d, t in sorted_keys:
                day_groups[d].append((d, t))

            rows_booth = []
            rows_roll = []

            for st in target_stations:
                d = self.station_data[st]
                factor_i = (1 + d.get('elect_rate',0)/100.0) * (1 + d['rate_intra']/100.0)
                factor_e = (1 + d['rate_extra']/100.0)

                time_map = temp_data[st]
                
                deltas_i = []
                deltas_e = []
                
                for day in day_groups:
                    prev_i = 0
                    prev_e = 0
                    for key in sorted(day_groups[day]):
                        if key not in time_map: continue
                        curr_i = time_map[key]['intra']
                        curr_e = time_map[key]['extra']
                        
                        d_i = max(0, curr_i - prev_i)
                        d_e = max(0, curr_e - prev_e)
                        deltas_i.append(d_i)
                        deltas_e.append(d_e)
                        
                        prev_i = curr_i
                        prev_e = curr_e
                
                st_disp = st.replace("사전투표소", "")

                # [기표대]
                if calc_booth:
                    peak_i = sum(sorted(deltas_i, reverse=True)[:3]) / 3 if deltas_i else 0
                    peak_e = sum(sorted(deltas_e, reverse=True)[:3]) / 3 if deltas_e else 0
                    
                    req_booth_i = max(2, math.ceil((peak_i * b_time_i) / 3600))
                    req_booth_e = max(2, math.ceil((peak_e * b_time_e) / 3600))
                    total_booths = req_booth_i + req_booth_e

                    rows_booth.append([
                        st_disp, 
                        total_booths,
                        int(peak_i), req_booth_i,
                        int(peak_e), req_booth_e
                    ])

                # [롤 용지]
                if calc_roll:
                    equip_i = self.station_data[st]['intra']
                    equip_e = self.station_data[st]['extra']
                    
                    dashboard_val_i = int(self.station_data[st]['past_intra'] * factor_i)
                    dashboard_val_e = int(self.station_data[st]['past_extra'] * factor_e)

                    avg_voter_i = dashboard_val_i / equip_i if equip_i > 0 else 0
                    avg_voter_e = dashboard_val_e / equip_e if equip_e > 0 else 0

                    pure_roll_i = max(1, math.ceil(avg_voter_i / r_cap_i)) * equip_i
                    pure_roll_e = max(1, math.ceil(avg_voter_e / r_cap_e)) * equip_e
                    
                    sub_total = pure_roll_i + pure_roll_e
                    reserve = math.ceil(sub_total * 0.1)
                    total_sum = sub_total + reserve

                    # [수정] 장비수 컬럼 추가 (관내/관외 각각)
                    rows_roll.append([
                        st_disp,
                        total_sum,
                        sub_total,
                        dashboard_val_i, 
                        equip_i,          # [신규] 관내 장비수
                        pure_roll_i,
                        dashboard_val_e, 
                        equip_e,          # [신규] 관외 장비수
                        pure_roll_e,
                        reserve
                    ])
            
            # --- 합계 행 추가 ---
            if calc_booth and rows_booth:
                sum_total_b = sum(r[1] for r in rows_booth)
                sum_intra_b = sum(r[3] for r in rows_booth)
                sum_extra_b = sum(r[5] for r in rows_booth)
                summary_booth = ["합계", sum_total_b, "", sum_intra_b, "", sum_extra_b]
                rows_booth.insert(0, summary_booth)

            if calc_roll and rows_roll:
                # 1. 일반 투표소 합계 계산 (장비수 포함)
                sum_total_r = sum(r[1] for r in rows_roll)
                sum_sub_r = sum(r[2] for r in rows_roll)
                
                sum_vote_i = sum(r[3] for r in rows_roll)
                sum_equip_i = sum(r[4] for r in rows_roll) # [신규]
                sum_roll_i = sum(r[5] for r in rows_roll)
                
                sum_vote_e = sum(r[6] for r in rows_roll)
                sum_equip_e = sum(r[7] for r in rows_roll) # [신규]
                sum_roll_e = sum(r[8] for r in rows_roll)
                
                sum_res = sum(r[9] for r in rows_roll)
                
                # 2. 위원회 보관분 계산 (투표소 예비용 총합의 20%)
                commission_keep = math.ceil(sum_res * 0.2)
                
                # [수정] 위원회 행 생성 로직 변경 (0 또는 빈칸 -> "-" 표시)
                # 구조: [이름, 합계, 소계, 관내3개, 관외3개, 예비용]
                comm_row = [
                    "위원회", 
                    commission_keep, # 합계 (보관분 자체)
                    "-",             # [변경] 소계: 0 -> "-"
                    "-", "-", "-",   # [변경] 관내 상세: 빈칸 -> "-"
                    "-", "-", "-",   # [변경] 관외 상세: 빈칸 -> "-"
                    commission_keep  # 예비용 컬럼에 값 표시
                ]
                
                # 4. 전체 합계 행 계산 (위원회 보관분 포함)
                final_grand_total = sum_total_r + commission_keep
                final_res_total = sum_res + commission_keep

                summary_roll = [
                    "합계", 
                    final_grand_total,
                    sum_sub_r, 
                    sum_vote_i, sum_equip_i, sum_roll_i, 
                    sum_vote_e, sum_equip_e, sum_roll_e, 
                    final_res_total
                ]
                
                # 5. 리스트에 삽입 (합계 -> 위원회 -> 투표소 순서)
                rows_roll.insert(0, comm_row)     # 2번째 줄 (인덱스 1이 되지만 일단 넣고)
                rows_roll.insert(0, summary_roll) # 1번째 줄 (맨 앞에 넣으므로 밀림)

            # --- 엑셀 저장 ---
            try:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                
                # [수정] 실행 파일 경로 확인
                if getattr(sys, 'frozen', False):
                    base_path = os.path.dirname(os.path.abspath(sys.executable))
                else:
                    base_path = os.path.dirname(os.path.abspath(__file__))
                
                # [수정] '임시저장' 폴더 생성 로직 추가
                save_dir = os.path.join(base_path, "임시저장")
                if not os.path.exists(save_dir):
                    os.makedirs(save_dir)
                    
                filename = f"물품소요량산출_{timestamp}.xlsx"
                save_path = os.path.join(save_dir, filename)
                
                wb = __import__('openpyxl').Workbook()
                if 'Sheet' in wb.sheetnames: wb.remove(wb['Sheet'])
                
                ws1 = None
                ws2 = None
                
                # ==================== 시트 1: 기표대 ====================
                if calc_booth:
                    ws1 = wb.create_sheet("기표대 소요량")
                    ws1['A1'] = "사전투표소명"
                    ws1['B1'] = "합계"
                    ws1['C1'] = f"관내({b_time_i}초)"
                    ws1['E1'] = f"관외({b_time_e}초)"
                    
                    ws1['C2'] = "최다투표자수\n(1시간)"
                    ws1['D2'] = "기표대"
                    ws1['E2'] = "최다투표자수\n(1시간)"
                    ws1['F2'] = "기표대"

                    ws1.merge_cells('A1:A2')
                    ws1.merge_cells('B1:B2')
                    ws1.merge_cells('C1:D1')
                    ws1.merge_cells('E1:F1')

                    for r in rows_booth:
                        ws1.append(r)
                    
                    last_row1 = ws1.max_row + 1
                    ws1.cell(row=last_row1, column=1).value = "*()는 선거인 1인의 투표 소요시간(발급시간 제외)을 말함."
                    ws1.merge_cells(start_row=last_row1, start_column=1, end_row=last_row1, end_column=6)
                    ws1.cell(row=last_row1, column=1).font = Font(size=9, italic=True)

                # ==================== 시트 2: 롤 용지 ====================
                if calc_roll:
                    ws2 = wb.create_sheet("롤 투표용지 소요량")
                    
                    # 1행 헤더
                    ws2['A1'] = "사전투표소명"
                    ws2['B1'] = "합계"
                    ws2['C1'] = "소계"
                    ws2['D1'] = f"관내({r_cap_i}명)" # D~F 병합 예정
                    ws2['G1'] = f"관외({r_cap_e}명)" # G~I 병합 예정
                    ws2['J1'] = "예비용"

                    # 2행 헤더 (장비수 추가됨)
                    ws2['D2'] = "예상투표자수"
                    ws2['E2'] = "장비수"      # [신규]
                    ws2['F2'] = "롤투표용지"
                    
                    ws2['G2'] = "예상투표자수"
                    ws2['H2'] = "장비수"      # [신규]
                    ws2['I2'] = "롤투표용지"

                    # 병합 설정
                    ws2.merge_cells('A1:A2')
                    ws2.merge_cells('B1:B2')
                    ws2.merge_cells('C1:C2')
                    ws2.merge_cells('D1:F1') # 관내 3칸
                    ws2.merge_cells('G1:I1') # 관외 3칸
                    ws2.merge_cells('J1:J2')

                    for r in rows_roll:
                        ws2.append(r)
                    
                    # 숫자 서식 적용 (J열까지)
                    for row in ws2.iter_rows(min_row=3, max_row=ws2.max_row, min_col=2, max_col=10):
                        for cell in row:
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0'

                    last_row2 = ws2.max_row + 1
                    ws2.cell(row=last_row2, column=1).value = "*()는 1롤 투표용지당 최대 사전투표자 수를 말함."
                    # 병합 범위 확장 (10열까지)
                    ws2.merge_cells(start_row=last_row2, start_column=1, end_row=last_row2, end_column=10)
                    ws2.cell(row=last_row2, column=1).font = Font(size=9, italic=True)

                # ==================== 스타일 적용 ====================
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
                header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                sum_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                
                def style_sheet(ws, limit_row):
                    max_col = ws.max_column
                    for row in ws.iter_rows(min_row=1, max_row=limit_row, min_col=1, max_col=max_col):
                        for cell in row:
                            cell.border = thin_border
                            cell.alignment = align_center
                            if cell.row <= 2:
                                cell.font = Font(bold=True)
                                cell.fill = header_fill
                            elif cell.row == 3:
                                cell.font = Font(bold=True)
                                cell.fill = sum_fill

                    ws.column_dimensions['A'].width = 18
                    for col in range(2, max_col + 1):
                        col_letter = get_column_letter(col)
                        ws.column_dimensions[col_letter].width = 13

                no_side = Side(border_style=None)
                no_border = Border(left=no_side, right=no_side, top=no_side, bottom=no_side)

                if calc_booth and ws1:
                    style_sheet(ws1, last_row1 - 1)
                    for col in range(1, 7):
                        ws1.cell(row=last_row1, column=col).border = no_border
                
                if calc_roll and ws2:
                    style_sheet(ws2, last_row2 - 1) # 전체 테두리 일단 적용

                    # 주석 행(맨 마지막) 테두리 제거 (J열까지)
                    for col in range(1, 11):
                        ws2.cell(row=last_row2, column=col).border = no_border

                wb.save(save_path)
                
                pop.destroy()
                if messagebox.askyesno("완료", f"파일이 생성되었습니다.\n\n{filename}\n\n지금 파일을 여시겠습니까?"):
                     if platform.system() == 'Windows':
                        os.startfile(save_path)
                        
            except Exception as e:
                import traceback
                traceback.print_exc()
                messagebox.showerror("저장 오류", f"파일 저장 중 오류가 발생했습니다.\n{e}", parent=pop)

        btn_run = ttk.Button(pop, text="💾 소요량 산출 및 엑셀 저장", command=_run_calculation)
        btn_run.pack(fill="x", padx=15, pady=20, ipady=5)

    def run_auto_balance(self, total_assets, total_reserve):
        self._ensure_data_loaded()
        
        # 1. 사용할 수 있는 실제 장비 수
        target_count = total_assets - total_reserve
        num_stations = len(self.station_data)
        
        # ---------------------------------------------------------
        # [수정 완료] 2. 기초 데이터 집계 (날짜+시간 구분 로직 적용)
        # ---------------------------------------------------------
        station_stats = {}

        # (1) 데이터를 날짜/시간 순서대로 정렬하기 위해 구조화
        # 구조: temp_data[투표소명][(day, time)] = {'intra':값, 'extra':값}
        temp_data = {}
        all_keys = set() # (day, time) 튜플을 저장

        for file in self.vote_files:
            if file not in self.cached_data: continue
            df, day, time = self.cached_data[file]
            
            # 시간이 없는 데이터는 제외
            if time is None: continue
            
            # [수정] 키를 (day, time)으로 설정하여 날짜별 중복 방지
            time_key = (day, time)
            all_keys.add(time_key)

            for idx, row in df.iterrows():
                st_name = str(row['사전투표소명']).strip()
                if st_name not in self.station_data: continue 
                
                if st_name not in temp_data:
                    temp_data[st_name] = {}
                
                # 예측 비율(가중치) 적용
                user_rate_intra = self.station_data[st_name]['rate_intra']
                user_rate_extra = self.station_data[st_name]['rate_extra']
                elect_rate = self.station_data[st_name].get('elect_rate', 0)
                
                factor_intra = (1 + (elect_rate / 100.0)) * (1 + (user_rate_intra / 100.0))
                factor_extra = (1 + (user_rate_extra / 100.0))
                
                try:
                    v_intra = float(row['관내사전투표자수']) * factor_intra
                    v_extra = float(row['관외사전투표자수']) * factor_extra
                    
                    temp_data[st_name][time_key] = {
                        'intra': v_intra,
                        'extra': v_extra
                    }
                except: pass
        
        # (2) 시간순으로 순회하며 '구간별 순증가분(Delta)' 계산
        sorted_keys = sorted(list(all_keys)) # [(1,6), (1,7)... (2,6), (2,7)...] 정렬됨
        
        # 날짜별로 그룹화 (날짜가 바뀌면 누적 계산을 초기화해야 하므로)
        from collections import defaultdict
        day_groups = defaultdict(list)
        for d, t in sorted_keys:
            day_groups[d].append((d, t))

        for st_name, time_map in temp_data.items():
            total_intra_in_target_time = 0
            total_extra_in_target_time = 0
            
            # 날짜별로 루프를 돌며 계산
            for day in day_groups:
                prev_intra = 0
                prev_extra = 0
                
                # 해당 날짜의 시간대들 순회 (예: 6시, 7시, ... 18시)
                for key in sorted(day_groups[day]):
                    if key not in time_map: continue
                    
                    day_val, time_val = key
                    
                    curr_intra = time_map[key]['intra']
                    curr_extra = time_map[key]['extra']
                    
                    # [핵심] 현재 누적값 - 이전 누적값 = 해당 시간대 순수 투표자 수
                    delta_intra = max(0, curr_intra - prev_intra)
                    delta_extra = max(0, curr_extra - prev_extra)
                    
                    # [필터링] 우리가 원하는 11시 ~ 18시 사이의 데이터만 합산
                    if 11 <= time_val <= 18:
                        total_intra_in_target_time += delta_intra
                        total_extra_in_target_time += delta_extra
                    
                    # 다음 루프를 위해 현재 값을 '이전 값'으로 갱신
                    prev_intra = curr_intra
                    prev_extra = curr_extra
            
            # 최종 계산된 값을 station_stats에 저장
            station_stats[st_name] = {
                'intra_voters': total_intra_in_target_time,
                'extra_voters': total_extra_in_target_time
            }
        # ---------------------------------------------------------
        
        # 3. 배분 알고리즘 시작
        # (1) 기본 할당: 모든 투표소의 관내/관외에 1대씩 강제 할당
        current_alloc = {}
        for st in self.station_data:
            current_alloc[st] = {'intra': 1, 'extra': 1}

        # 남은 장비(remaining) 계산
        remaining = target_count - (num_stations * 2)
        
        # (2) Greedy Algorithm
        while remaining > 0:
            max_load = -1 
            target_info = None 
            
            for st in current_alloc:
                # [안전장치] 해당 투표소의 데이터가 없으면 0 처리
                if st not in station_stats:
                    s_stat = {'intra_voters': 0, 'extra_voters': 0}
                else:
                    s_stat = station_stats[st]

                weight_extra = 1.18

                # 관내 부하 계산
                curr_intra = current_alloc[st]['intra']
                if curr_intra > 0:
                    load_intra = s_stat['intra_voters'] / curr_intra
                else:
                    load_intra = float('inf')

                if load_intra > max_load:
                    max_load = load_intra
                    target_info = (st, 'intra')
                
                # 관외 부하 계산
                curr_extra = current_alloc[st]['extra']
                if curr_extra > 0:
                    load_extra = (s_stat['extra_voters'] * weight_extra) / curr_extra
                else:
                    load_extra = float('inf')

                if load_extra > max_load:
                    max_load = load_extra
                    target_info = (st, 'extra')
            
            if target_info:
                st_name, r_type = target_info
                current_alloc[st_name][r_type] += 1
                remaining -= 1
            else:
                break

        # 4. 결과 집계 및 UI 반영
        total_intra_used = 0
        total_extra_used = 0
        
        for item_id in self.tree.get_children():
            item_values = self.tree.item(item_id)['values']
            if not item_values: continue
            
            # [수정] 화면에 보이는 이름(values[0]) 대신 고유 ID(item_id) 사용
            # item_id에는 '진잠동사전투표소' 같은 풀네임이 들어있음
            st_name = item_id 
            
            if st_name in self.station_data:
                new_intra = current_alloc[st_name]['intra']
                new_extra = current_alloc[st_name]['extra']
                
                # 데이터 저장
                self.station_data[st_name]['intra'] = new_intra
                self.station_data[st_name]['extra'] = new_extra
                
                total_intra_used += new_intra
                total_extra_used += new_extra
                
                # UI 업데이트용 변수
                org_intra = self.station_data[st_name]['org_intra']
                org_extra = self.station_data[st_name]['org_extra']
                val_rate_intra = self.station_data[st_name]['rate_intra']
                val_rate_extra = self.station_data[st_name]['rate_extra']
                elect_disp = item_values[1]

                disp_intra = f"{org_intra} → {new_intra}" if new_intra != org_intra else str(new_intra)
                disp_extra = f"{org_extra} → {new_extra}" if new_extra != org_extra else str(new_extra)
                rate_txt = self._get_merged_rate_text(val_rate_intra, val_rate_extra)
                
                st_disp = st_name.replace("사전투표소", "") # [추가]
                self.tree.item(item_id, values=(st_disp, elect_disp, disp_intra, disp_extra, rate_txt))
        
        # 5. 결과 메시지
        final_used = total_intra_used + total_extra_used
        msg = (f"배분 완료!\n\n"
               f"■ 총 보유 장비: {total_assets}대\n"
               f"■ 실제 배치: {final_used}대 (관내 {total_intra_used} / 관외 {total_extra_used})\n"
               f"■ 예비 장비: {total_reserve}대")
               
        self.log(f"[자동 배분] 총 {total_assets}대 중 {final_used}대 배치 완료. (예비 {total_reserve})")
        messagebox.showinfo("배분 완료", msg)
        
    def _open_rate_input_dialog(self, st_name, item_id, elect_disp, curr_intra, curr_extra, org_intra, org_extra):
        # 현재 값 가져오기
        cur_r_intra = self.station_data[st_name]['rate_intra']
        cur_r_extra = self.station_data[st_name]['rate_extra']

        # 팝업창 생성
        pop = tk.Toplevel(self.root)
        pop.title("조정률 개별 설정")
        
        # [수정] 안내 문구가 들어갈 공간 확보를 위해 높이를 180 -> 220으로 변경
        pop.geometry("260x220")
        pop.resizable(False, False)
        
        # 중앙 배치
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - 130
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - 110 # 높이 변경 반영
        pop.geometry(f"+{x}+{y}")

        ttk.Label(pop, text=f"[{st_name}]", font=("맑은 고딕", 10, "bold")).pack(pady=(15, 5))

        # [추가] 안내 문구 라벨 (문구 수정)
        guide_msg = "※ 이 설정은 투표율이 아닌\n사전투표자 수의 증가율(%)입니다."
        ttk.Label(pop, text=guide_msg, justify="center", foreground="blue", font=("맑은 고딕", 8)).pack(pady=(0, 10))

        frame_in = ttk.Frame(pop)
        frame_in.pack(fill="x", padx=30, pady=5)
        ttk.Label(frame_in, text="관내 조정(%):").pack(side="left")
        # (수정) f-string 포맷팅(.1f) 적용
        entry_intra = ttk.Entry(frame_in, width=10, justify="right")
        entry_intra.insert(0, f"{cur_r_intra:.1f}")  # <-- 소수점 1자리로 고정하여 표시
        entry_intra.pack(side="right")

        frame_out = ttk.Frame(pop)
        frame_out.pack(fill="x", padx=30, pady=5)
        ttk.Label(frame_out, text="관외 조정(%):").pack(side="left")
        entry_extra = ttk.Entry(frame_out, width=10, justify="right")
        entry_extra.insert(0, f"{cur_r_extra:.1f}")  # <-- 소수점 1자리로 고정하여 표시
        entry_extra.pack(side="right")

        def _apply():
            try:
                # [수정] 정수(int) 대신 실수(float)로 받아서 소수점 입력 가능하게 변경
                new_r_intra = float(entry_intra.get())
                new_r_extra = float(entry_extra.get())
                
                # 데이터 업데이트
                self.station_data[st_name]['rate_intra'] = new_r_intra
                self.station_data[st_name]['rate_extra'] = new_r_extra
                
                # 화면 갱신
                rate_txt = self._get_merged_rate_text(new_r_intra, new_r_extra)
                disp_intra = f"{org_intra} → {curr_intra}" if curr_intra != org_intra else str(curr_intra)
                disp_extra = f"{org_extra} → {curr_extra}" if curr_extra != org_extra else str(curr_extra)

                st_disp = st_name.replace("사전투표소", "")
                self.tree.item(item_id, values=(st_disp, elect_disp, disp_intra, disp_extra, rate_txt))
                
                # [핵심 추가] 전체 통계 재계산 (Bottom-Up 방식)
                self.recalculate_grand_total()
                
                self.log(f"{st_name} 조정률 변경: 내 {new_r_intra}% / 외 {new_r_extra}%")
                pop.destroy()
            except ValueError:
                messagebox.showerror("오류", "숫자만 입력해주세요.", parent=pop)

        ttk.Button(pop, text="적용", command=_apply).pack(pady=15, fill='x', padx=30)

    # [수정] 인자에 cmap='Greens', title_suffix='' 추가
    def _plot_page(self, df, scenarios, stations_list, filename=None, is_pdf=False, cmap='Greens', title_suffix=''):
        count = len(scenarios)
        
        # 1. 기본 단위 높이 계산
        if is_pdf:
            unit_h = 13 
        else:
            unit_h = max(7, 4 + (len(stations_list) * 0.6))

        # 2. 행/열 및 전체 크기 자동 계산
        if count == 1: 
            nrows, ncols = 1, 1
            figsize = (12, unit_h)
        elif count == 2: 
            nrows, ncols = 1, 2
            figsize = (20, unit_h)
        elif count <= 4: 
            nrows, ncols = 2, 2
            figsize = (20, unit_h * 2) 
        else: 
            nrows, ncols = 3, 2
            figsize = (20, unit_h * 3) 

        # 3. 서브플롯 생성
        fig, axes = plt.subplots(nrows, ncols, figsize=figsize)
        
        if count == 1: axes_flat = [axes]
        else: axes_flat = axes.flatten()

        # === [핵심 수정 1] 최대값(vmax) 계산 로직 ===
        global_max = 1
        
        if '투표자' in title_suffix:
            # [유지] 투표자 수 모드는 데이터 내 최대값 기준 (상대평가)
            temp_max = 0
            for _, _, _, v_col, _, _, active in scenarios:
                if active and v_col in df.columns:
                    current_max = df[v_col].max() 
                    if current_max > temp_max:
                        temp_max = current_max
            global_max = temp_max if temp_max > 0 else 100
        else:
            # [수정] 혼잡도 모드는 절대값 100 기준 고정
            global_max = 100

        for idx, (day, type_name, label_col, value_col, eq_col, org_eq_col, _) in enumerate(scenarios):
            ax = axes_flat[idx]
            
            if str(day) == '전체':
                df_day = df[df['일차'] == '전체']
            else:
                df_day = df[df['일차'] == day]
            
            if df_day.empty:
                ax.text(0.5, 0.5, '데이터 없음', ha='center', va='center')
                continue
            
            pivot = df_day.pivot_table(index=label_col, columns='시간대', values=value_col)
            
            # === [수정] 모드에 따라 첫 번째 열(행 통계) 계산 방식 변경 ===
            avg_label = '' 
            
            if '투표자' in title_suffix:
                # [투표자수 모드]: 합계(Sum)
                pivot[avg_label] = pivot.sum(axis=1)
            else:
                # [혼잡도 모드]: 평균(Mean)
                pivot[avg_label] = pivot.mean(axis=1)

            # [수정] 아래쪽(열) 통계도 모드에 따라 '합계' 또는 '평균'으로 변경
            if '투표자' in title_suffix:
                avg_row = pivot.sum(axis=0)
            else:
                avg_row = pivot.mean(axis=0)
                
            pivot.loc[avg_label] = avg_row
            
            # 정렬 및 재배치
            time_cols = sorted([c for c in pivot.columns if c != avg_label])
            new_cols = [avg_label] + time_cols
            pivot = pivot[new_cols]
            
            target_labels = [s.replace('사전투표소','') for s in stations_list]
            valid_labels = [l for l in target_labels if l in pivot.index]
            new_rows = [avg_label] + valid_labels
            pivot = pivot.reindex(new_rows)

            # 장비 데이터 준비
            equip_data = df_day.drop_duplicates(subset=[label_col]).set_index(label_col)[[eq_col, org_eq_col]]

            annot_labels = []
            for row_label in new_rows:
                if row_label == avg_label:
                    annot_labels.append("") 
                else:
                    try:
                        curr = int(equip_data.loc[row_label, eq_col])
                        org = int(equip_data.loc[row_label, org_eq_col])
                        if curr != org: txt = f"{org} → {curr}"
                        else: txt = f"{curr}"
                        annot_labels.append(txt)
                    except: annot_labels.append("?")

            equip_df = pd.DataFrame(1, index=new_rows, columns=['장비']) 
            equip_df.iloc[0] = 0 

            annot_matrix = pd.DataFrame(annot_labels, index=new_rows, columns=['장비'])

            divider = make_axes_locatable(ax)
            ax_equip = divider.append_axes("left", size="7%", pad=0.08) 
            
            custom_cmap = ListedColormap(['white', '#F0F4F8'])

            sns.heatmap(equip_df, annot=annot_matrix, fmt='', 
                        cmap=custom_cmap, vmin=0, vmax=1,
                        cbar=False, xticklabels=False,
                        linewidths=0.5, linecolor='white', ax=ax_equip)
            
            ax_equip.set_xlabel("")
            ax_equip.set_ylabel("사전투표소", fontsize=11, fontweight='bold')
            ax_equip.tick_params(axis='y', rotation=0, length=0)

            ax_equip.text(0.5, 0.95, "장비수", ha='center', va='bottom', fontsize=10, fontweight='bold', color='black')
            
            # [수정] 라벨 텍스트
            col_label_txt = "시간대별 합계 →" if '투표자' in title_suffix else "시간대별 평균 →"
            ax_equip.text(0.95, 0.5, col_label_txt, ha='right', va='center', fontsize=9, fontweight='bold', color='#3B5BDB')

            # 주석(Annotation) 포맷
            target_hours = [c for c in pivot.columns if isinstance(c, (int, float)) and 11 <= c <= 18]

            if '투표자' in title_suffix:
                annot_df = pivot.applymap(lambda x: f"{x:,.0f}")
            else:
                annot_df = pivot.applymap(lambda x: f"{x:.1f}")
                if target_hours:
                    mean_11_18 = pivot.loc[avg_label, target_hours].mean()
                    original_text = annot_df.iloc[0, 0]
                    annot_df.iloc[0, 0] = f"{original_text}\n({mean_11_18:.1f})"

            # === [핵심 수정 2] 색상 정규화 (자기들끼리 비교) ===
            pivot_color = pivot.copy()
            
            if '투표자' in title_suffix:
                # [논리] 합계 열/행을 0으로 만드는 대신, 본문의 진하기(global_max)에 맞춰 비율을 조정(Scaling)함
                
                # 1. 투표소별 합계 (첫번째 열, Grand Total 제외)
                # 이 열에서 가장 큰 값을 찾아서, 그 값이 global_max(가장 진한 색)가 되도록 비율 조정
                col_data = pivot.iloc[1:, 0]
                if not col_data.empty and col_data.max() > 0:
                    scaled_col = (col_data / col_data.max()) * global_max
                    pivot_color.iloc[1:, 0] = scaled_col

                # 2. 시간대별 합계 (첫번째 행, Grand Total 제외)
                # 이 행에서 가장 큰 값을 찾아서, 그 값이 global_max가 되도록 비율 조정
                row_data = pivot.iloc[0, 1:]
                if not row_data.empty and row_data.max() > 0:
                    scaled_row = (row_data / row_data.max()) * global_max
                    pivot_color.iloc[0, 1:] = scaled_row

                # 3. 전체 합계 (좌측 상단, Grand Total)
                # 이 값은 무조건 가장 크므로 가장 진한 색(global_max)으로 고정
                pivot_color.iloc[0, 0] = global_max
            
            # [수정] annot(글자)에는 원래 숫자(annot_df)가 들어가고, 색상(data)에는 조정된 값(pivot_color)이 들어감
            sns.heatmap(pivot_color, annot=annot_df, fmt='', cmap=cmap, cbar=False, 
                        linewidths=0.5, linecolor='white', vmin=0, vmax=global_max, ax=ax)
            
            if '투표자' in title_suffix:
                row_label_txt = "↓ 투표소별\n합계" 
            else:
                row_label_txt = "↓ 투표소별\n평균"
                
            ax.text(0.5, -0.2, row_label_txt, ha='center', va='bottom', fontsize=10, fontweight='bold', color='#3B5BDB', clip_on=False)
            
            # 테두리 그리기
            rect_row = patches.Rectangle((0, 0), len(pivot.columns), 1, linewidth=3, edgecolor='#3B5BDB', facecolor='none', clip_on=False)
            ax.add_patch(rect_row)
            rect_col = patches.Rectangle((0, 0), 1, len(pivot), linewidth=3, edgecolor='#3B5BDB', facecolor='none', clip_on=False)
            ax.add_patch(rect_col)

            ax.set_ylabel("") 
            ax.set_yticks([]) 
            ax.xaxis.tick_top()
            ax.xaxis.set_label_position('top')
            ax.tick_params(axis='x', length=0)

            if time_cols:
                try:
                    times = [int(c) for c in time_cols]
                    start_t, end_t = min(times), max(times)
                    expected_count = end_t - start_t + 1
                    
                    if len(times) != expected_count:
                        raise ValueError(f"데이터 오류: 시간대 불연속")
                    labels = [''] + list(range(start_t, end_t + 1))
                except Exception:
                    labels = [''] + time_cols
            else:
                labels = ['']

            ticks = np.arange(len(pivot.columns)) + 0.5
            ax.set_xticks(ticks)
            ax.set_xticklabels(labels, rotation=0)

            if str(day) == '전체':
                day_str = "전체(합계)" if '투표자' in title_suffix else "전체(평균)"
            else:
                day_str = f"{day}일차"

            title_txt = f"{type_name} 사전투표 ({day_str}) - {title_suffix}"
            
            ax.set_title(title_txt, fontsize=14, fontweight='bold', pad=20)
            ax.set_xlabel('시간대', fontsize=11, fontweight='bold')

        for i in range(count, len(axes_flat)):
            axes_flat[i].axis('off')

        # [수정] 지역명 대괄호 제거 및 텍스트 정리 (대전광역시 띄어쓰기 추가)
        clean_region = self.region_name.replace('[', '').replace(']', '').strip() if self.region_name else ""
        # '대전광역시' 뒤에 공백을 추가하고, 혹시 모를 이중 공백을 하나로 정리합니다.
        clean_region = clean_region.replace("대전광역시", "대전광역시 ").replace("  ", " ").strip()
        
        if '투표자' in title_suffix:
            # 투표자 수 모드 (제목 변경 요청 반영)
            main_text = f"{clean_region} 사전투표소 예상 혼잡도"
            sub_text = "(단위: 명)"
        else:
            # [요청하신 부분] 발급능력 모드
            main_text = f"{clean_region} 사전투표 발급자 현황"
            sub_text = "(장비 1대 기준)"

        # 1. 메인 제목 (크게) - y 위치를 조금 조정하여 공간 확보
        fig.suptitle(main_text, fontsize=22, fontweight='bold', y=0.98)
        
        # 2. 부제 (작게) - 메인 제목 바로 아래에 위치
        fig.text(0.5, 0.94, sub_text, ha='center', va='top', fontsize=12, color='#555555')
        
        if cmap == 'Greens':
            desc_text = "※ 각 셀의 수치: 장비 1대당 1시간 평균 처리 인원 (혼잡도)"
            legend_text = "테두리: 전체 시간 평균  |  ( 괄호 안 ): 11~18시 집중평균  |  장비: [기존] → [변경]"
        else:
            desc_text = "※ 각 셀의 수치: 해당 시간대의 실제 투표자 수 합계 (단위: 명)"
            legend_text = "테두리: 전체 시간 합계  |  장비: [기존] → [변경]"
        
        fig.text(0.5, 0.02, 
                    f"{desc_text}\n{legend_text}", 
                    ha='center', fontsize=11, fontweight='bold', color='#333333')
        
        plt.tight_layout(rect=[0, 0.05, 1, 0.95]) 
        
        if filename and not is_pdf:
            plt.savefig(filename)
            plt.close(fig)
            
        return fig

    def create_dashboard_ui(self, parent):
        pnl = ttk.LabelFrame(parent, text=" 📊 사전투표율 시뮬레이션 ", padding="10")
        pnl.pack(fill="x", pady=(0, 10))

        # 1. 직전 선거 정보 (첫 번째 줄)
        self.lbl_past_info = ttk.Label(pnl, text="직전 사전투표율: - % (총 -명 / 관내 -명 / 관외 -명)", font=("맑은 고딕", 9))
        self.lbl_past_info.pack(anchor="w", pady=(0, 5))

        # 2. 당해 선거 예상 정보 (두 번째 줄) - 수정 가능하도록 분리
        f_pred = ttk.Frame(pnl)
        f_pred.pack(fill="x", pady=(0, 10))
        
        # 앞부분 라벨
        ttk.Label(f_pred, text="예상 사전투표율: ", font=("맑은 고딕", 9, "bold"), foreground="blue").pack(side="left")
        
        # 중간 입력창 (숫자 수정 가능)
        self.entry_predict_rate = ttk.Entry(f_pred, width=8, justify="right", font=("맑은 고딕", 9, "bold"), foreground="blue")
        self.entry_predict_rate.pack(side="left")
        self.entry_predict_rate.bind("<Return>", self._on_predict_rate_confirm) # 엔터키 바인딩
        
        # 뒷부분 상세 정보 라벨
        self.lbl_predict_details = ttk.Label(f_pred, text="% (총 -명 / 관내 -명 / 관외 -명)", 
                                             font=("맑은 고딕", 9, "bold"), foreground="blue")
        self.lbl_predict_details.pack(side="left")

        # 3. 슬라이더 컨트롤 (세 번째 줄)
        f_ctrl = ttk.Frame(pnl)
        f_ctrl.pack(fill="x")
        
        ttk.Label(f_ctrl, text="사전투표자 증가율: ").pack(side="left") # <--- 변경됨
        
        self.var_rate = tk.DoubleVar(value=0.0)
        # [수정] 범위를 0 ~ 100으로 변경 (음수 불가)
        self.scale_rate = ttk.Scale(f_ctrl, from_=0, to=100, variable=self.var_rate, command=self.on_slider_drag)
        self.scale_rate.pack(side="left", fill="x", expand=True, padx=5)
        
        # 증감률 표시용 입력창
        self.entry_rate = ttk.Entry(f_ctrl, width=6, justify="right", font=("맑은 고딕", 9))
        self.entry_rate.pack(side="left", padx=(5, 0))
        self.entry_rate.insert(0, "0.0")
        self.entry_rate.bind("<Return>", self._on_entry_rate_confirm) 
        
        ttk.Label(f_ctrl, text="%").pack(side="left", padx=(2, 0))

        # 0% 초기화 버튼 (목표 설정 버튼은 제거함)
        btn_reset_rate = ttk.Button(f_ctrl, text="↺ 0%", width=6, command=self.reset_rate_zero)
        btn_reset_rate.pack(side="left", padx=(5, 0))

    def _update_dashboard_info(self):
        # 1. 직전 선거 정보 업데이트
        if self.total_past_electors > 0:
            rate = (self.total_past_voters / self.total_past_electors) * 100
            
            msg = (f"직전 사전투표율: {rate:.2f}% "
                   f"(총 {self.total_past_voters:,}명 / 관내 {self.total_past_intra:,}명 / 관외 {self.total_past_extra:,}명)")
            self.lbl_past_info.config(text=msg)
            
            # 2. [수정] 슬라이더를 무조건 0%로 고정 (사용자 요청)
            # 선거인수가 늘어도 투표자 수는 그대로 시작 -> 투표율은 과거보다 낮게 나옴
            self.reset_rate_zero()
        else:
            # 과거 데이터가 없으면 0으로 초기화
            self.reset_rate_zero()
    
    def reset_rate_zero(self):
        # [기능] 슬라이더와 증감률을 즉시 0으로 초기화
        self.var_rate.set(0.0)
        self.entry_rate.delete(0, tk.END)
        self.entry_rate.insert(0, "0.0")
        self.on_slider_drag(0.0)
    
    def _on_predict_rate_confirm(self, event):
        if self.total_past_voters == 0 or self.total_recent_electors == 0:
            messagebox.showwarning("데이터 부족", "기초 데이터가 로드되지 않았습니다.")
            return

        try:
            target_turnout = float(self.entry_predict_rate.get())
            target_voters = self.total_recent_electors * (target_turnout / 100.0)
            
            # [수정 정밀화] 역산 로직: 개별 투표소의 인구 변동률을 합산하여 정확한 베이스라인 산출
            baseline_voters = 0
            for st_name, data in self.station_data.items():
                p_intra = data.get('past_intra', 0)
                p_extra = data.get('past_extra', 0)
                e_rate = data.get('elect_rate', 0)
                
                # [수정] 반올림(round) 적용
                baseline_voters += int(round(p_intra * (1 + e_rate / 100.0))) + int(round(p_extra))
            
            if baseline_voters == 0: return

            # 필요 증감률(R) 계산: Target = Baseline * (1 + R/100)
            # 1 + R/100 = Target / Baseline
            # R = (Target / Baseline - 1) * 100
            required_rate = ((target_voters / baseline_voters) - 1) * 100
            
            # [수정] 범위 제한 로직 변경 (0% ~ 100%)
            if required_rate > 100:
                required_rate = 100
                messagebox.showwarning("범위 제한", "최대 증가율(100%)을 초과하여 100%로 설정합니다.")
            elif required_rate < 0:
                required_rate = 0
                messagebox.showwarning("범위 제한", "설정된 투표율이 너무 낮습니다.\n증가율은 0% 미만(감소)으로 설정할 수 없습니다.")
            
            self.root.focus()
            self.var_rate.set(required_rate)
            self.entry_rate.delete(0, tk.END)
            self.entry_rate.insert(0, f"{required_rate:.1f}")
            
            self.on_slider_drag(required_rate)
            
        except ValueError:
            self.on_slider_drag(self.var_rate.get())

    def on_slider_drag(self, val):
        try:
            rate = float(val)
        except:
            rate = 0.0
            
        # 1. 입력창 텍스트 갱신
        if self.root.focus_get() != self.entry_rate:
            self.entry_rate.delete(0, tk.END)
            self.entry_rate.insert(0, f"{rate:.1f}")

        # 2. 예상 인원 및 투표율 계산
        # [수정 핵심] 전체 평균 비율 대신, 개별 투표소의 변동분을 합산하여 정확도 향상 (엑셀 리포트와 일치시킴)
        slider_factor = 1 + (rate / 100.0)
        
        pred_intra = 0
        pred_extra = 0
        
        # 개별 투표소 데이터를 순회하며 정밀 합산
        for st_name, data in self.station_data.items():
            p_intra = data.get('past_intra', 0)
            p_extra = data.get('past_extra', 0)
            e_rate = data.get('elect_rate', 0) # 선거인수 변동률
            
            # [관내]
            val_intra = p_intra * (1 + e_rate / 100.0) * slider_factor
            
            # [관외]
            val_extra = p_extra * slider_factor
            
            # [수정 핵심] 단순 버림(int) 대신 반올림(round)을 사용하여 총합의 통계적 오차 감소
            pred_intra += int(round(val_intra))
            pred_extra += int(round(val_extra))
            
        pred_total = pred_intra + pred_extra
        
        pred_rate = 0.0
        if self.total_recent_electors > 0:
            pred_rate = (pred_total / self.total_recent_electors) * 100
            
        # 3. UI 업데이트
        if self.root.focus_get() != self.entry_predict_rate:
            self.entry_predict_rate.delete(0, tk.END)
            self.entry_predict_rate.insert(0, f"{pred_rate:.2f}")
            
        details = f"% (총 {int(pred_total):,}명 / 관내 {int(pred_intra):,}명 / 관외 {int(pred_extra):,}명)"
        if hasattr(self, 'lbl_predict_details'):
            self.lbl_predict_details.config(text=details)
            
        # 4. 트리뷰 리스트 업데이트
        self.update_treeview_by_rate(rate)

    def _on_entry_rate_confirm(self, event):
        # [기능] 입력창에 직접 숫자를 쓰고 엔터를 쳤을 때
        try:
            val = float(self.entry_rate.get())
            
            # [수정] 슬라이더 범위(0 ~ 100) 제한
            if val > 100: val = 100
            elif val < 0: val = 0
            
            # 슬라이더 위치 이동
            self.var_rate.set(val)
            
            # [수정] 포커스를 메인 윈도우(root)로 옮겨서 입력창에서 커서를 뺌
            self.root.focus() 
            
            # 값 적용 실행
            self.on_slider_drag(val)
            
        except ValueError:
            # 숫자가 아닌 값을 입력하면 0으로 초기화
            self.reset_rate_zero()

    def update_treeview_by_rate(self, val):
        full_rate = float(val)       # [수정 1] 계산용 정밀 값 (대시보드와 동일)
        disp_rate = round(full_rate, 1) # [수정 2] 화면 표시용 반올림 값
        
        for item_id in self.tree.get_children():
            st_name = item_id 
            if st_name in self.station_data:
                # [핵심 수정] 데이터에는 '정밀 값(full_rate)'을 저장하여 롤 용지 계산 시 대시보드와 일치시킴
                self.station_data[st_name]['rate_intra'] = full_rate
                self.station_data[st_name]['rate_extra'] = full_rate
                
                # 화면 갱신용 데이터 준비
                elect_disp = self.tree.item(item_id)['values'][1]
                curr_intra = self.station_data[st_name]['intra']
                curr_extra = self.station_data[st_name]['extra']
                org_intra = self.station_data[st_name]['org_intra']
                org_extra = self.station_data[st_name]['org_extra']
                
                disp_intra = f"{org_intra} → {curr_intra}" if curr_intra != org_intra else str(curr_intra)
                disp_extra = f"{org_extra} → {curr_extra}" if curr_extra != org_extra else str(curr_extra)

                # [수정 3] 화면 텍스트(rate_txt)는 보기 좋게 반올림된 값 사용
                rate_txt = self._get_merged_rate_text(disp_rate, disp_rate)
                st_disp = st_name.replace("사전투표소", "")

                self.tree.item(item_id, values=(st_disp, elect_disp, disp_intra, disp_extra, rate_txt))

    def recalculate_grand_total(self):
        # [기능] 개별 투표소의 설정을 집계하여 전체 통계(상단 UI) 역업데이트
        
        total_exp_voters = 0
        total_exp_intra = 0
        total_exp_extra = 0
        
        # 1. 모든 투표소 순회하며 예상 인원 합산
        for st_name, data in self.station_data.items():
            p_intra = data.get('past_intra', 0)
            p_extra = data.get('past_extra', 0)
            r_intra = data.get('rate_intra', 0)
            r_extra = data.get('rate_extra', 0)
            
            # [수정] 선거인수 변동률(elect_rate)도 함께 반영해야 정확한 예상이 됨
            e_rate = data.get('elect_rate', 0)

            # 복합 증감률 공식: (1 + 인구증감) * (1 + 사용자조정)
            factor_i = (1 + e_rate / 100.0) * (1 + r_intra / 100.0)
            factor_e = (1 + r_extra / 100.0) # 관외는 인구증감 영향 없음(기존 로직 따름)

            exp_i = p_intra * factor_i
            exp_e = p_extra * factor_e
            
            # [수정 핵심] 소수점 처리 방식 통일! (엑셀/슬라이더와 동일하게 int로 변환 후 합산)
            # 이렇게 해야 개별 수정 시에도 전체 합계가 정확히 맞아떨어집니다.
            val_i = int(round(exp_i))
            val_e = int(round(exp_e))
            
            total_exp_voters += (val_i + val_e)
            total_exp_intra += val_i
            total_exp_extra += val_e
            
        # 2. UI 업데이트 (입력창 + 상세 라벨)
        if self.total_recent_electors > 0:
            new_turnout = (total_exp_voters / self.total_recent_electors) * 100
            
            # 예상 사전투표율 입력창 업데이트
            self.entry_predict_rate.delete(0, tk.END)
            self.entry_predict_rate.insert(0, f"{new_turnout:.2f}")
            
            # 상세 정보 라벨 업데이트
            details = f"% (총 {int(total_exp_voters):,}명 / 관내 {int(total_exp_intra):,}명 / 관외 {int(total_exp_extra):,}명)"
            self.lbl_predict_details.config(text=details)

            # (3) 슬라이더 위치 업데이트 (평균 증감률로 표시)
            # [수정] 인구증가분이 중복 반영되지 않도록, '인구 변동만 반영된 베이스라인'을 구해서 역산
            baseline_sum = 0
            for st_name, data in self.station_data.items():
                p_intra = data.get('past_intra', 0)
                p_extra = data.get('past_extra', 0)
                e_rate = data.get('elect_rate', 0)
                
                # 슬라이더(심리적 증가)가 0%일 때의 자연스러운 예상치 합산
                # 관내: 인구변동 반영 / 관외: 인구변동 미반영(기존로직 유지)
                baseline_sum += int(round(p_intra * (1 + e_rate / 100.0))) + int(round(p_extra))

            if baseline_sum > 0:
                # (현재 설정된 총합 / 베이스라인 - 1) * 100 = 순수 심리적 증가율
                avg_rate = ((total_exp_voters / baseline_sum) - 1) * 100
                
                self.var_rate.set(avg_rate)
                
                # [수정됨] 슬라이더 옆 입력창 갱신
                self.entry_rate.delete(0, tk.END)
                self.entry_rate.insert(0, f"{avg_rate:.1f}")
            else:
                self.var_rate.set(0.0)

if __name__ == "__main__":
    root = tk.Tk()
    app = ElectionAnalyzerApp(root)
    root.mainloop()
